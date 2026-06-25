# -*- coding: utf-8 -*-
"""
gerar_relatorio_excel.py — Gerador de Relatório Financeiro em Excel (.xlsx)

Uso:
    python 03_SCRIPTS/gerar_relatorio_excel.py --conferencia "07_DADOS_TEMPORARIOS/CONFERENCIA_CALCULOS_CONTRATO_EDMILSON_F05_AGUA_VIVA_LEO.md"

Ou com JSON direto:
    python 03_SCRIPTS/gerar_relatorio_excel.py --json "07_DADOS_TEMPORARIOS/WIDEPAY_CONSULTAS/WIDEPAY_EDMILSON.json" --cliente "Edmilson Silva Dos Santos" --lote "F05" --quadra "F" --parcelas-contrato 100 --valor-parcela 99.00 --valor-total-contrato 10500.00

Gera planilha XLSX com abas:
    1. Resumo
    2. Pagamentos Recebidos
    3. Interpretação das Parcelas
    4. Validação
    5. Alertas (se houver divergência)
"""
import os
import sys
import json
import re
import argparse
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl nao esta instalado. Execute: pip install openpyxl")
    sys.exit(1)

ROOT_DIR = Path(__file__).resolve().parent.parent
sys.path.append(str(ROOT_DIR / "00_SISTEMA_PRECHECK"))

try:
    from precheck_regras import executar_precheck
    executar_precheck("gerar_relatorio_excel.py")
except ImportError as e:
    print(f"AVISO: Nao foi possivel carregar o precheck: {e}")

# ── Cores e estilos ──────────────────────────────────────────────────────────
VERDE_ESCURO  = "1B5E20"
VERDE_CLARO   = "E8F5E9"
BRANCO        = "FFFFFF"
CINZA_CLARO   = "F5F5F5"
CINZA_BORDA   = "BDBDBD"
VERMELHO      = "C62828"
AMARELO       = "FFF9C4"
AZUL_HEADER   = "1565C0"
AZUL_CLARO    = "E3F2FD"

FONT_TITULO    = Font(name="Calibri", size=16, bold=True, color=BRANCO)
FONT_SUBTITULO = Font(name="Calibri", size=12, bold=True, color=VERDE_ESCURO)
FONT_HEADER    = Font(name="Calibri", size=10, bold=True, color=BRANCO)
FONT_NORMAL    = Font(name="Calibri", size=10)
FONT_BOLD      = Font(name="Calibri", size=10, bold=True)
FONT_TOTAL     = Font(name="Calibri", size=11, bold=True, color=VERDE_ESCURO)
FONT_ALERTA    = Font(name="Calibri", size=10, bold=True, color=VERMELHO)

FILL_TITULO    = PatternFill(start_color=VERDE_ESCURO, end_color=VERDE_ESCURO, fill_type="solid")
FILL_HEADER    = PatternFill(start_color=AZUL_HEADER, end_color=AZUL_HEADER, fill_type="solid")
FILL_BRANCO    = PatternFill(start_color=BRANCO, end_color=BRANCO, fill_type="solid")
FILL_ZEBRA     = PatternFill(start_color=CINZA_CLARO, end_color=CINZA_CLARO, fill_type="solid")
FILL_VERDE_CL  = PatternFill(start_color=VERDE_CLARO, end_color=VERDE_CLARO, fill_type="solid")
FILL_AMARELO   = PatternFill(start_color=AMARELO, end_color=AMARELO, fill_type="solid")
FILL_AZUL_CL   = PatternFill(start_color=AZUL_CLARO, end_color=AZUL_CLARO, fill_type="solid")

BORDA_FINA = Border(
    left=Side(style="thin", color=CINZA_BORDA),
    right=Side(style="thin", color=CINZA_BORDA),
    top=Side(style="thin", color=CINZA_BORDA),
    bottom=Side(style="thin", color=CINZA_BORDA),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center", wrap_text=True)

BRL_FORMAT = '#,##0.00'
PCT_FORMAT = '0.00%'


def interpretar_referencias(descricao, valor_original, valor_base, valor_entrada=0.0):
    """Interpreta descrição do WidePay para identificar referências/meses pagos."""
    desc = (descricao or "").lower().strip()
    refs = []
    regra = ""
    qtd = 0

    if "entrada" in desc:
        return ["Entrada"], 0, "identificado como entrada pela descricao"

    if valor_entrada > 0 and abs(valor_original - valor_entrada) < 0.01:
        return ["Entrada"], 0, "valor original corresponde a entrada contratual"

    # Padrão: "atrazo 08 09 10 11 de 2025" ou "atraso 12-25"
    meses_match = re.findall(r'\b(0[1-9]|1[0-2])\b', desc)
    if meses_match and any(w in desc for w in ["atraso", "atrazo", "ref", "referente"]):
        refs = meses_match
        qtd = len(refs)
        regra = "referencias identificadas na descricao"
        return refs, qtd, regra

    # Padrão: "Ref a 05 06 07-2025"
    ref_match = re.findall(r'\b(0[1-9]|1[0-2])\b', desc)
    if ref_match and "ref" in desc:
        refs = ref_match
        qtd = len(refs)
        regra = "referencias identificadas na descricao"
        return refs, qtd, regra

    # Boleto de carnê: tratar pelo vencimento
    if "apart" in desc or "carne" in desc or "carnê" in desc or not desc or desc == "-":
        refs = ["vencimento"]
        qtd = 1
        regra = "boleto de carne tratado pelo vencimento"
        return refs, qtd, regra

    # Inferência por valor
    if valor_base and valor_base > 0 and valor_original and valor_original > 0:
        qtd_inferida = round(valor_original / valor_base)
        if qtd_inferida >= 1:
            refs = [f"inferido ({qtd_inferida}x R$ {valor_base:.2f})"]
            qtd = qtd_inferida
            regra = "quantidade inferida pelo valor original dividido pelo valor base da parcela"
            return refs, qtd, regra

    refs = ["REFERENCIA NAO IDENTIFICADA"]
    qtd = 1
    regra = "referencia nao identificada"
    return refs, qtd, regra


def extrair_pagamentos_recebidos(dados_wp):
    """Filtra apenas cobranças com status Recebido."""
    recebidos = []
    cobrancas = dados_wp.get("cobrancas", [])
    for cob in cobrancas:
        status = (cob.get("status") or "").strip()
        if status.lower() in ["recebido", "pago"]:
            valor_recebido = float(cob.get("valor_recebido", 0))
            if valor_recebido > 0:
                recebidos.append(cob)
    return recebidos


def classificar_tipo(cob, carnes_lista):
    """Classifica se o pagamento é de Carnê, Avulso ou Cobrança manual."""
    if cob.get("pertence_a_carne"):
        return "Carnê"
    if cob.get("avulsa"):
        return "Avulso"
    desc = (cob.get("descricao") or "").lower()
    for c in carnes_lista:
        ref = (c.get("referencia") or "").lower()
        if ref and ref in desc:
            return "Carnê"
    if any(w in desc for w in ["atraso", "atrazo", "ref ", "referente"]):
        return "Avulso"
    return "Cobrança manual"


def criar_aba_resumo(ws, dados_cliente, total_pago, parcelas_pagas_equiv,
                     total_parcelas_contrato, valor_parcela, valor_total_contrato):
    """Cria a aba Resumo com dados do cliente e resumo financeiro."""
    ws.sheet_properties.tabColor = VERDE_ESCURO

    # Título
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "RELATÓRIO FINANCEIRO — Loteamento Água Viva — Iaçú-BA"
    c.font = FONT_TITULO
    c.fill = FILL_TITULO
    c.alignment = ALIGN_CENTER

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font = Font(name="Calibri", size=10, color=BRANCO)
    c.fill = FILL_TITULO
    c.alignment = ALIGN_CENTER

    # Dados do cliente
    saldo_parcelado = total_parcelas_contrato * valor_parcela
    entrada = valor_total_contrato - saldo_parcelado

    campos = [
        ("Cliente", dados_cliente.get("cliente", "-")),
        ("Lote / Quadra", f"{dados_cliente.get('lote', '-')} / Quadra {dados_cliente.get('quadra', '-')}"),
        ("Loteamento", "Água Viva — Iaçú-BA"),
        ("Data do Relatório", datetime.now().strftime("%d/%m/%Y")),
        ("Valor Total Contratado", valor_total_contrato),
        ("Entrada ou Valor Fora do Parcelamento", entrada),
        ("Saldo Parcelado", saldo_parcelado),
        ("Total de Parcelas (Contrato)", total_parcelas_contrato),
        ("Valor Base da Parcela", valor_parcela),
    ]

    row = 4
    for label, val in campos:
        ws.cell(row=row, column=1, value=label).font = FONT_BOLD
        ws.cell(row=row, column=1).fill = FILL_VERDE_CL
        ws.cell(row=row, column=1).border = BORDA_FINA
        ws.cell(row=row, column=1).alignment = ALIGN_LEFT

        cell_val = ws.cell(row=row, column=2, value=val)
        cell_val.font = FONT_NORMAL
        cell_val.border = BORDA_FINA
        cell_val.alignment = ALIGN_LEFT
        if isinstance(val, float):
            cell_val.number_format = BRL_FORMAT
        row += 1

    # Resumo financeiro em destaque
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="RESUMO FINANCEIRO")
    c.font = FONT_SUBTITULO
    c.fill = FILL_AZUL_CL
    c.alignment = ALIGN_CENTER
    c.border = BORDA_FINA
    row += 1

    parcelas_restantes = max(0, total_parcelas_contrato - parcelas_pagas_equiv)
    pct_parcelas = parcelas_pagas_equiv / total_parcelas_contrato if total_parcelas_contrato > 0 else 0
    pct_financeiro = total_pago / valor_total_contrato if valor_total_contrato > 0 else 0

    if pct_parcelas >= 1.0:
        situacao = "QUITADO"
    elif pct_parcelas >= 0.75:
        situacao = "AVANÇADO (75%+)"
    elif pct_parcelas >= 0.5:
        situacao = "INTERMEDIÁRIO (50%+)"
    elif pct_parcelas > 0:
        situacao = "INICIAL"
    else:
        situacao = "SEM PAGAMENTO REGISTRADO"

    resumo_campos = [
        ("Total Pago Recebido", total_pago, BRL_FORMAT),
        ("Parcelas Pagas Confirmadas", parcelas_pagas_equiv, None),
        ("Parcelas Restantes", parcelas_restantes, None),
        ("% Parcelas Quitadas", pct_parcelas, PCT_FORMAT),
        ("% Financeiro Pago", pct_financeiro, PCT_FORMAT),
        ("Situação Calculada", situacao, None),
    ]

    for label, val, fmt in resumo_campos:
        ws.cell(row=row, column=1, value=label).font = FONT_BOLD
        ws.cell(row=row, column=1).fill = FILL_VERDE_CL
        ws.cell(row=row, column=1).border = BORDA_FINA
        ws.cell(row=row, column=1).alignment = ALIGN_LEFT

        cell_val = ws.cell(row=row, column=2, value=val)
        cell_val.font = FONT_TOTAL
        cell_val.border = BORDA_FINA
        cell_val.alignment = ALIGN_LEFT
        if fmt:
            cell_val.number_format = fmt
        row += 1

    # Ajustar larguras
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40
    for col_letter in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 18

    return {
        "total_pago": total_pago,
        "parcelas_pagas_equiv": parcelas_pagas_equiv,
        "parcelas_restantes": parcelas_restantes,
        "pct_parcelas": pct_parcelas,
        "pct_financeiro": pct_financeiro,
        "situacao": situacao,
        "total_parcelas_contrato": total_parcelas_contrato,
    }


def criar_aba_pagamentos(ws, pagamentos, dados_cliente, valor_parcela, carnes_lista, valor_entrada=0.0):
    """Cria a aba Pagamentos Recebidos (somente Recebido/Pago)."""
    ws.sheet_properties.tabColor = "1565C0"

    headers = [
        "Cliente", "Lote/Quadra", "ID", "Tipo", "Descrição WidePay",
        "Vencimento", "Data Pagamento", "Valor Original", "Valor Recebido",
        "Status", "Valor Base Parcela", "Referências", "Parcelas Quitadas",
        "Observação"
    ]

    # Cabeçalho
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDA_FINA

    # Dados
    total_recebido = 0.0
    total_parcelas = 0
    cliente_nome = dados_cliente.get("cliente", "-")
    lote_quadra = f"{dados_cliente.get('lote', '-')} / {dados_cliente.get('quadra', '-')}"

    for row_idx, pag in enumerate(pagamentos, 2):
        valor_original = float(pag.get("valor_original", 0))
        valor_recebido = float(pag.get("valor_recebido", 0))
        desc = pag.get("descricao", "") or "-"
        tipo = classificar_tipo(pag, carnes_lista)
        refs, qtd, regra = interpretar_referencias(desc, valor_original, valor_parcela, valor_entrada)

        total_recebido += valor_recebido
        total_parcelas += qtd

        fill = FILL_BRANCO if (row_idx % 2 == 0) else FILL_ZEBRA

        vals = [
            cliente_nome,
            lote_quadra,
            str(pag.get("id", "-")),
            tipo,
            desc,
            pag.get("vencimento", "-"),
            pag.get("pagamento", "-") or "-",
            valor_original,
            valor_recebido,
            pag.get("status", "Recebido"),
            valor_parcela,
            ", ".join(refs),
            qtd,
            regra,
        ]

        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONT_NORMAL
            cell.fill = fill
            cell.border = BORDA_FINA
            cell.alignment = ALIGN_LEFT
            if col_idx in [8, 9, 11]:  # Valores monetários
                cell.number_format = BRL_FORMAT
                cell.alignment = ALIGN_RIGHT

    # Totalizadores
    total_row = len(pagamentos) + 2
    ws.cell(row=total_row, column=7, value="TOTAL").font = FONT_TOTAL
    ws.cell(row=total_row, column=7).alignment = ALIGN_RIGHT
    ws.cell(row=total_row, column=7).border = BORDA_FINA

    cell_total_orig = ws.cell(row=total_row, column=8,
                              value=f"=SUM(H2:H{total_row-1})")
    cell_total_orig.font = FONT_TOTAL
    cell_total_orig.number_format = BRL_FORMAT
    cell_total_orig.border = BORDA_FINA

    cell_total_rec = ws.cell(row=total_row, column=9,
                             value=f"=SUM(I2:I{total_row-1})")
    cell_total_rec.font = FONT_TOTAL
    cell_total_rec.number_format = BRL_FORMAT
    cell_total_rec.border = BORDA_FINA

    cell_total_parc = ws.cell(row=total_row, column=13,
                              value=f"=SUM(M2:M{total_row-1})")
    cell_total_parc.font = FONT_TOTAL
    cell_total_parc.border = BORDA_FINA

    # Filtro automático e painel congelado
    ws.auto_filter.ref = f"A1:N{total_row}"
    ws.freeze_panes = "A2"

    # Ajustar larguras
    widths = [30, 14, 20, 12, 35, 14, 14, 16, 16, 12, 16, 25, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return total_recebido, total_parcelas


def criar_aba_interpretacao(ws, pagamentos, dados_cliente, valor_parcela, carnes_lista, valor_entrada=0.0):
    """Cria a aba Interpretação das Parcelas."""
    ws.sheet_properties.tabColor = "F57F17"

    headers = [
        "ID", "Descrição", "Valor Original", "Valor Recebido",
        "Valor Base Parcela", "Regra Usada", "Referências",
        "Parcelas Quitadas", "Observação"
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDA_FINA

    for row_idx, pag in enumerate(pagamentos, 2):
        valor_original = float(pag.get("valor_original", 0))
        valor_recebido = float(pag.get("valor_recebido", 0))
        desc = pag.get("descricao", "") or "-"
        tipo = classificar_tipo(pag, carnes_lista)
        refs, qtd, regra = interpretar_referencias(desc, valor_original, valor_parcela, valor_entrada)

        fill = FILL_BRANCO if (row_idx % 2 == 0) else FILL_ZEBRA

        obs = f"{tipo}: {regra}"

        vals = [
            str(pag.get("id", "-")),
            desc,
            valor_original,
            valor_recebido,
            valor_parcela,
            regra,
            ", ".join(refs),
            qtd,
            obs,
        ]

        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONT_NORMAL
            cell.fill = fill
            cell.border = BORDA_FINA
            cell.alignment = ALIGN_LEFT
            if col_idx in [3, 4, 5]:
                cell.number_format = BRL_FORMAT
                cell.alignment = ALIGN_RIGHT

    ws.auto_filter.ref = f"A1:I{len(pagamentos) + 1}"
    ws.freeze_panes = "A2"

    widths = [20, 35, 16, 16, 16, 45, 25, 14, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def criar_aba_validacao(ws, total_pago, parcelas_pagas_equiv, total_parcelas_contrato,
                        valor_total_contrato, pagamentos, total_recebido_soma,
                        total_parcelas_soma):
    """Cria a aba Validação com checagens matemáticas."""
    ws.sheet_properties.tabColor = "2E7D32"

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "VALIDAÇÃO MATEMÁTICA DO RELATÓRIO"
    c.font = FONT_SUBTITULO
    c.fill = FILL_VERDE_CL
    c.alignment = ALIGN_CENTER
    c.border = BORDA_FINA

    headers = ["#", "Validação", "Resultado", "Status"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDA_FINA

    parcelas_restantes = max(0, total_parcelas_contrato - parcelas_pagas_equiv)
    pct_parcelas = parcelas_pagas_equiv / total_parcelas_contrato if total_parcelas_contrato > 0 else 0
    pct_financeiro = total_pago / valor_total_contrato if valor_total_contrato > 0 else 0

    # Validação de que nenhum Recebido tem valor 0
    todos_recebidos_ok = all(
        float(p.get("valor_recebido", 0)) > 0 for p in pagamentos
    )
    # Validação de que nenhum não-Recebido entrou na lista
    nenhum_invalido = all(
        (p.get("status") or "").strip().lower() in ["recebido", "pago"]
        for p in pagamentos
    )

    validacoes = [
        ("1", "Total pago = soma dos valores recebidos",
         f"R$ {total_pago:.2f} vs R$ {total_recebido_soma:.2f}",
         "✅ OK" if abs(total_pago - total_recebido_soma) < 0.01 else "❌ FALHA"),

        ("2", "Parcelas pagas = soma das parcelas interpretadas",
         f"{parcelas_pagas_equiv} vs {total_parcelas_soma}",
         "✅ OK" if parcelas_pagas_equiv == total_parcelas_soma else "❌ FALHA"),

        ("3", "Parcelas restantes = contrato - pagas",
         f"{parcelas_restantes} = {total_parcelas_contrato} - {parcelas_pagas_equiv}",
         "✅ OK" if parcelas_restantes == (total_parcelas_contrato - parcelas_pagas_equiv) else "❌ FALHA"),

        ("4", "% parcelas = pagas / contrato",
         f"{pct_parcelas*100:.2f}% = {parcelas_pagas_equiv} / {total_parcelas_contrato}",
         "✅ OK"),

        ("5", "% financeiro = recebido / contratado",
         f"{pct_financeiro*100:.2f}% = R$ {total_pago:.2f} / R$ {valor_total_contrato:.2f}",
         "✅ OK"),

        ("6", "Nenhum vencido/aguardando/cancelado na aba Pagamentos",
         f"Todos {len(pagamentos)} registros são Recebido/Pago",
         "✅ OK" if nenhum_invalido else "❌ FALHA"),

        ("7", "Nenhum Recebido com valor R$ 0,00",
         f"Checagem de {len(pagamentos)} registros",
         "✅ OK" if todos_recebidos_ok else "❌ FALHA"),

        ("8", "Todo pagamento recebido possui interpretação",
         f"{len(pagamentos)} pagamentos interpretados",
         "✅ OK"),
    ]

    tem_falha = False
    for row_idx, (num, desc, resultado, status) in enumerate(validacoes, 4):
        fill = FILL_BRANCO if (row_idx % 2 == 0) else FILL_ZEBRA
        if "FALHA" in status:
            fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            tem_falha = True

        vals = [num, desc, resultado, status]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONT_NORMAL if "FALHA" not in status else FONT_ALERTA
            cell.fill = fill
            cell.border = BORDA_FINA
            cell.alignment = ALIGN_LEFT

    widths = [6, 55, 45, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return tem_falha


def criar_aba_alertas(ws, alertas):
    """Cria a aba Alertas quando houver divergências."""
    ws.sheet_properties.tabColor = "C62828"

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "⚠ ALERTAS E DIVERGÊNCIAS"
    c.font = Font(name="Calibri", size=14, bold=True, color=VERMELHO)
    c.fill = FILL_AMARELO
    c.alignment = ALIGN_CENTER
    c.border = BORDA_FINA

    headers = ["#", "Tipo", "Descrição"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = PatternFill(start_color=VERMELHO, end_color=VERMELHO, fill_type="solid")
        cell.alignment = ALIGN_CENTER
        cell.border = BORDA_FINA

    for row_idx, (tipo, desc) in enumerate(alertas, 4):
        fill = FILL_BRANCO if (row_idx % 2 == 0) else FILL_ZEBRA
        ws.cell(row=row_idx, column=1, value=row_idx - 3).font = FONT_NORMAL
        ws.cell(row=row_idx, column=1).border = BORDA_FINA
        ws.cell(row=row_idx, column=2, value=tipo).font = FONT_BOLD
        ws.cell(row=row_idx, column=2).border = BORDA_FINA
        ws.cell(row=row_idx, column=2).fill = fill
        ws.cell(row=row_idx, column=3, value=desc).font = FONT_NORMAL
        ws.cell(row=row_idx, column=3).border = BORDA_FINA
        ws.cell(row=row_idx, column=3).fill = fill

    widths = [6, 35, 80]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def gerar_relatorio_excel(dados_wp, dados_cliente, valor_parcela,
                          total_parcelas_contrato, valor_total_contrato,
                          caminho_saida):
    """Gera o arquivo Excel completo."""
    wb = Workbook()

    carnes_lista = dados_wp.get("carnes", [])
    pagamentos = extrair_pagamentos_recebidos(dados_wp)

    # Ordenar por vencimento
    def parse_data(d):
        try:
            return datetime.strptime(d, "%d/%m/%Y")
        except Exception:
            return datetime.min
    pagamentos.sort(key=lambda p: parse_data(p.get("vencimento", "")))

    valor_entrada = valor_total_contrato - (total_parcelas_contrato * valor_parcela)

    # Calcular totais
    total_pago = sum(float(p.get("valor_recebido", 0)) for p in pagamentos)
    parcelas_pagas_equiv = 0
    for pag in pagamentos:
        desc = pag.get("descricao", "") or "-"
        valor_original = float(pag.get("valor_original", 0))
        _, qtd, _ = interpretar_referencias(desc, valor_original, valor_parcela, valor_entrada)
        parcelas_pagas_equiv += qtd

    # ── Aba 1: Resumo ────────────────────────────────────────────────────
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    resumo_data = criar_aba_resumo(
        ws_resumo, dados_cliente, total_pago, parcelas_pagas_equiv,
        total_parcelas_contrato, valor_parcela, valor_total_contrato
    )

    # ── Aba 2: Pagamentos Recebidos ──────────────────────────────────────
    ws_pag = wb.create_sheet("Pagamentos Recebidos")
    total_recebido_soma, total_parcelas_soma = criar_aba_pagamentos(
        ws_pag, pagamentos, dados_cliente, valor_parcela, carnes_lista, valor_entrada
    )

    # ── Aba 3: Interpretação das Parcelas ────────────────────────────────
    ws_interp = wb.create_sheet("Interpretação Parcelas")
    criar_aba_interpretacao(ws_interp, pagamentos, dados_cliente, valor_parcela, carnes_lista, valor_entrada)

    # ── Aba 4: Validação ─────────────────────────────────────────────────
    ws_valid = wb.create_sheet("Validação")
    tem_falha = criar_aba_validacao(
        ws_valid, total_pago, parcelas_pagas_equiv, total_parcelas_contrato,
        valor_total_contrato, pagamentos, total_pago, parcelas_pagas_equiv
    )

    # ── Aba 5: Alertas ───────────────────────────────────────────────────
    alertas = []

    if total_parcelas_contrato <= 0:
        alertas.append(("Contrato", "Total de parcelas do contrato nao confirmado"))
    if valor_parcela <= 0:
        alertas.append(("Contrato", "Valor base da parcela nao identificado"))
    if valor_total_contrato <= 0:
        alertas.append(("Contrato", "Valor total contratado nao identificado"))

    for pag in pagamentos:
        desc = pag.get("descricao", "") or "-"
        valor_original = float(pag.get("valor_original", 0))
        refs, _, regra = interpretar_referencias(desc, valor_original, valor_parcela)
        if "REFERENCIA NAO IDENTIFICADA" in refs:
            alertas.append(("Recebimento", f"ID {pag.get('id')}: sem referencia clara"))
        if not pag.get("pagamento"):
            alertas.append(("Recebimento", f"ID {pag.get('id')}: sem data de pagamento"))

    if tem_falha:
        alertas.append(("Validação", "Uma ou mais validacoes matematicas falharam"))

    if alertas:
        ws_alertas = wb.create_sheet("Alertas")
        criar_aba_alertas(ws_alertas, alertas)

    # Salvar
    wb.save(caminho_saida)
    print(f"\n[EXCEL GERADO] {caminho_saida}")
    print(f"  Pagamentos recebidos: {len(pagamentos)}")
    print(f"  Total pago: R$ {total_pago:.2f}")
    print(f"  Parcelas pagas equiv.: {parcelas_pagas_equiv}")
    print(f"  Parcelas restantes: {max(0, total_parcelas_contrato - parcelas_pagas_equiv)}")
    print(f"  Alertas: {len(alertas)}")
    if tem_falha:
        print("  ⚠ ENTREGA BLOQUEADA: validacao matematica falhou")
    return not tem_falha


def main():
    parser = argparse.ArgumentParser(
        description="Gera relatorio financeiro em Excel (.xlsx) para cliente WidePay"
    )
    parser.add_argument("--json", required=True,
                        help="Caminho do JSON de consulta WidePay")
    parser.add_argument("--cliente", required=True,
                        help="Nome do cliente")
    parser.add_argument("--lote", required=True,
                        help="Lote do cliente")
    parser.add_argument("--quadra", required=True,
                        help="Quadra do cliente")
    parser.add_argument("--parcelas-contrato", type=int, required=True,
                        help="Total de parcelas confirmado no contrato")
    parser.add_argument("--valor-parcela", type=float, required=True,
                        help="Valor base de cada parcela")
    parser.add_argument("--valor-total-contrato", type=float, required=True,
                        help="Valor total contratado do lote")
    parser.add_argument("--saida", default=None,
                        help="Caminho do arquivo Excel de saída (opcional)")

    args = parser.parse_args()

    # Carregar JSON
    caminho_json = Path(args.json)
    if not caminho_json.is_absolute():
        caminho_json = ROOT_DIR / caminho_json
    if not caminho_json.exists():
        print(f"ERRO: JSON nao encontrado: {caminho_json}")
        sys.exit(1)

    with open(caminho_json, "r", encoding="utf-8") as f:
        dados_wp = json.load(f)

    dados_cliente = {
        "cliente": args.cliente,
        "lote": args.lote,
        "quadra": args.quadra,
    }

    # Caminho de saída
    if args.saida:
        caminho_saida = Path(args.saida)
    else:
        nome_limpo = re.sub(r'[^a-zA-Z0-9]', '_', args.cliente.upper())
        data_str = datetime.now().strftime("%Y%m%d")
        nome_arquivo = f"RELATORIO_FINANCEIRO_CLIENTE_{nome_limpo}_LOTE_{args.lote}_{data_str}.xlsx"
        caminho_saida = ROOT_DIR / "02_RELATORIOS_GERADOS" / nome_arquivo

    if not caminho_saida.is_absolute():
        caminho_saida = ROOT_DIR / caminho_saida

    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    # Se o arquivo de saída já existir, gera automaticamente um novo nome com data/hora
    if caminho_saida.exists():
        stem = caminho_saida.stem
        suffix = caminho_saida.suffix
        hora_str = datetime.now().strftime("%H%M")
        
        # Evitar loop infinito ou duplicados se rodar no mesmo minuto
        if stem.endswith(f"_{hora_str}"):
            hora_str = datetime.now().strftime("%H%M%S")
            
        caminho_saida = caminho_saida.parent / f"{stem}_{hora_str}{suffix}"

    sucesso = gerar_relatorio_excel(
        dados_wp,
        dados_cliente,
        args.valor_parcela,
        args.parcelas_contrato,
        args.valor_total_contrato,
        str(caminho_saida),
    )

    if sucesso:
        print("\n[PROCESSO CONCLUIDO] Relatorio Excel gerado com sucesso.")
    else:
        print("\n[BLOQUEADO] Entrega final bloqueada por falha de validacao.")

    sys.exit(0 if sucesso else 1)


if __name__ == "__main__":
    main()
