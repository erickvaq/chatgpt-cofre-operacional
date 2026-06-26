# -*- coding: utf-8 -*-
import os
import sys
import argparse
import asyncio
from datetime import datetime
from pathlib import Path

# Adicionar a pasta raiz e a pasta do app ao sys.path
ROOT_DIR = Path(__file__).resolve().parent.parent
sys.path.append(str(ROOT_DIR))
sys.path.append(str(ROOT_DIR / "WideAPP_EXTRA"))

# Precheck de regras (opcional em modo portátil)
PRECHECK_DIR = ROOT_DIR / "00_SISTEMA_PRECHECK"
if PRECHECK_DIR.exists():
    sys.path.append(str(PRECHECK_DIR))
    try:
        from precheck_regras import executar_precheck
        executar_precheck("executar_auditoria.py")
        print("OK: precheck de regras executado")
    except Exception as e:
        print(f"Erro ao carregar o precheck de regras: {e}")
        sys.exit(1)
else:
    print("AVISO: 00_SISTEMA_PRECHECK nao encontrado — modo portavel, precheck ignorado")

from app.login_navegador import garantir_navegador_conectado
from app.extrator_widepay import extrair_dados_cliente
from app.leitor_contratos import carregar_dados_contrato, localizar_diretorio_cliente, PASTAS_PROJETO
from app.normalizador_pagamentos import normalizar_e_deduplicar
from app.calculadora_financeira import calcular_resumo
from app.validador_matematico import validar_conciliacao
from app.gerador_relatorios import exportar_relatorios_finais
from app.rastreabilidade import registrar_log, preparar_diretorio_entrega

def obter_todos_clientes_cadastrados():
    import re
    clientes = set()
    base_dir = PASTAS_PROJETO.get("Agua Viva")
    if base_dir and base_dir.exists():
        for quadra_dir in os.listdir(base_dir):
            quadra_path = base_dir / quadra_dir
            if quadra_path.is_dir() and "quadra" in quadra_dir.lower():
                for cliente_dir in os.listdir(quadra_path):
                    cliente_path = quadra_path / cliente_dir
                    if cliente_path.is_dir():
                        # Limpeza básica do lote do nome da pasta para obter o nome do cliente
                        nome = re.sub(r'\b[A-H]\d+\b', '', cliente_dir, flags=re.IGNORECASE)
                        nome = re.sub(r'\b(Lote|Quadra|Agua Viva|Leandro Meirelles|carne\d*|apart\d*|wp-pdf-\w+|v\d+|final|corrigido|previa)\b', '', nome, flags=re.IGNORECASE)
                        nome = re.sub(r'\b(docx|pdf|txt|html|md|jpg|jpeg|png)\b', '', nome, flags=re.IGNORECASE)
                        nome = nome.replace("-", " ").replace("_", " ").strip()
                        nome = re.sub(r'\s+', ' ', nome)
                        if len(nome) > 3:
                            clientes.add((nome.title(), cliente_dir))
    return sorted(list(clientes))

async def auditar_cliente_unico(ws_url, cliente_nome, lote_opcao=None):
    """Executa o pipeline completo de auditoria para um único cliente."""
    print(f"\n=======================================================")
    print(f" AUDITANDO CLIENTE: {cliente_nome.upper()}")
    print(f"=======================================================")
    
    try:
        registrar_log("INFO", "INICIANDO_AUDITORIA", cliente_nome, f"Lote: {lote_opcao or '-'}")
        
        # 1. Carregar dados do contrato local
        dados_contrato = carregar_dados_contrato(cliente_nome, lote_opcao)
        if not dados_contrato:
            registrar_log("ERRO", "FALHA_CONTRATO", cliente_nome, "Contrato local nao localizado.")
            dados_contrato = {
                'cliente': cliente_nome, 'lote': lote_opcao or '-', 'quadra': '-',
                'valor_parcela': 0.0, 'total_parcelas': 0, 'valor_total_contrato': 0.0
            }
            
        lote_final = dados_contrato.get("lote") or lote_opcao or "-"
        
        nome_busca = dados_contrato.get("cliente") or cliente_nome
        # 2. Extrair dados reais do WidePay via CDP
        dados_raw_wp = await extrair_dados_cliente(ws_url, nome_busca, lote_final, dados_contrato.get("quadra"))
        if dados_raw_wp.get("cliente"):
            dados_contrato["cliente"] = dados_raw_wp["cliente"]
        
        # 3. Normalizar e de-duplicar lançamentos
        valor_base = float(dados_contrato.get("valor_parcela") or 0.0)
        dados_normalizados = normalizar_e_deduplicar(dados_raw_wp, valor_base)
        
        # 4. Reconciliação financeira e cálculos
        dados_calculados = calcular_resumo(dados_normalizados, dados_contrato)
        
        # 5. Validação matemática de regras
        status_val, notas, bloqueado = validar_conciliacao(dados_calculados, dados_contrato)
        
        print(f"Status da Auditoria: {status_val}")
        for nota in notas:
            print(f"- {nota}")
            
        # 6. Gerar relatórios finais
        lote_entrega = lote_opcao if (lote_opcao and lote_opcao != "-") else lote_final
        pasta_entrega = preparar_diretorio_entrega(cliente_nome, lote_entrega)
        data_sufixo = datetime.now().strftime("%Y%m%d_%H%M")
        
        dados_contrato_copia = dados_contrato.copy()
        dados_contrato_copia["lote"] = lote_entrega
        
        arquivos = exportar_relatorios_finais(dados_contrato_copia, dados_calculados, dados_normalizados, pasta_entrega, data_sufixo)
        
        registrar_log("SUCESSO", "AUDITORIA_CONCLUIDA", cliente_nome, f"Status: {status_val}; Arquivos gerados na pasta de entrega.")
        
        return {
            "cliente": cliente_nome,
            "lote": lote_entrega,
            "status": status_val,
            "bloqueado": bloqueado,
            "notas": "; ".join(notas),
            "caminho_entrega": str(pasta_entrega),
            "arquivos": arquivos
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        registrar_log("ERRO", "FALHA_AUDITORIA", cliente_nome, str(e))
        print(f"Erro na auditoria do cliente {cliente_nome}: {e}")
        return None

async def main_async():
    parser = argparse.ArgumentParser(description="Auditoria Financeira WidePay Inteligente")
    parser.add_argument("--cliente", help="Nome do cliente especifico")
    parser.add_argument("--clientes", help="Lista de clientes separados por virgula (ex: 'Edmilson,Ana,Jose')")
    parser.add_argument("--lote", help="Lote especifico para desambiguacao")
    parser.add_argument("--lotes", help="Lista de lotes correspondentes aos clientes, separados por virgula")
    parser.add_argument("--letra", help="Inicial do nome do cliente")
    parser.add_argument("--letra-fim", help="Letra final para intervalo (se --letra for usada)")
    parser.add_argument("--quadra", help="Filtrar por uma quadra especifica")
    parser.add_argument("--todos", action="store_true", help="Processar todos os clientes cadastrados")
    parser.add_argument("--consolidado", action="store_true", help="Gera um relatorio consolidado unico")
    
    args = parser.parse_args()
    
    # 1. Garantir conexao com navegador e aba do WidePay
    print("Iniciando navegador...")
    try:
        ws_url = garantir_navegador_conectado()
    except Exception as e:
        print(f"Erro ao conectar com o navegador dedicado: {e}")
        sys.exit(1)
        
    # 2. Obter lista de clientes cadastrados fisicamente no projeto
    clientes_cadastrados = obter_todos_clientes_cadastrados()
    clientes_a_processar = []
    
    if args.cliente:
        # Um único cliente (retrocompatível)
        clientes_a_processar.append((args.cliente, args.lote))
    elif args.clientes:
        # Lista CSV de clientes: "Edmilson,Ana,Jose"
        nomes = [n.strip() for n in args.clientes.split(",") if n.strip()]
        lotes = [l.strip() for l in args.lotes.split(",") if l.strip()] if args.lotes else []
        for i, nome in enumerate(nomes):
            lote = lotes[i] if i < len(lotes) else None
            clientes_a_processar.append((nome, lote))
    elif args.letra:
        letra_ini = args.letra.lower()
        letra_fim = args.letra_fim.lower() if args.letra_fim else letra_ini
        for nome, pasta in clientes_cadastrados:
            inicial = nome.lower()[0]
            if letra_ini <= inicial <= letra_fim:
                clientes_a_processar.append((nome, None))
    elif args.quadra:
        quadra_alvo = args.quadra.lower()
        for nome, pasta in clientes_cadastrados:
            if f"quadra {quadra_alvo}" in pasta.lower() or f"_{quadra_alvo}" in pasta.lower() or pasta.lower().endswith(quadra_alvo):
                clientes_a_processar.append((nome, None))
    elif args.todos:
        for nome, pasta in clientes_cadastrados:
            clientes_a_processar.append((nome, None))
    else:
        print("Erro: Especifique pelo menos um parametro de selecao (--cliente, --clientes, --letra, --quadra ou --todos).")
        sys.exit(1)
        
    if not clientes_a_processar:
        print("Nenhum cliente selecionado para processamento.")
        sys.exit(0)
        
    print(f"Total de clientes a processar: {len(clientes_a_processar)}")
    
    # 3. Agrupar em blocos de até 3 clientes
    tamanho_bloco = 3
    blocos = [clientes_a_processar[i:i + tamanho_bloco] for i in range(0, len(clientes_a_processar), tamanho_bloco)]
    
    resultados = []
    from app.extrator_widepay import extrair_dados_clientes_bloco
    
    for bloco_idx, bloco in enumerate(blocos, start=1):
        print(f"\n=======================================================")
        print(f" PROCESSANDO BLOCO {bloco_idx}/{len(blocos)} ({len(bloco)} clientes)")
        print(f"=======================================================")
        
        bloco_info = []
        for cliente_nome, lote_opcao in bloco:
            # Carregar dados do contrato local
            dados_contrato = carregar_dados_contrato(cliente_nome, lote_opcao)
            if not dados_contrato:
                registrar_log("ERRO", "FALHA_CONTRATO", cliente_nome, "Contrato local nao localizado.")
                dados_contrato = {
                    'cliente': cliente_nome, 'lote': lote_opcao or '-', 'quadra': '-',
                    'valor_parcela': 0.0, 'total_parcelas': 0, 'valor_total_contrato': 0.0
                }
            lote_final = dados_contrato.get("lote") or lote_opcao or "-"
            nome_busca = dados_contrato.get("cliente") or cliente_nome
            
            bloco_info.append({
                "nome": nome_busca,
                "lote": lote_final,
                "lote_opcao": lote_opcao,
                "quadra": dados_contrato.get("quadra") or "-",
                "dados_contrato": dados_contrato
            })
            
        try:
            # Executa a extração em bloco no WidePay
            resultados_raw = await extrair_dados_clientes_bloco(ws_url, bloco_info)
            
            # Processar os resultados de cada cliente do bloco
            for cli in bloco_info:
                cliente_nome = cli["nome"]
                dados_contrato = cli["dados_contrato"]
                lote_final = cli["lote"]
                lote_opcao = cli.get("lote_opcao")
                print(f"[DEBUG_SCOPES] cli_lote_opcao={cli.get('lote_opcao')}, lote_opcao_val={lote_opcao}")
                
                print(f"\n--- Processando auditoria individual: {cliente_nome} (Lote: {lote_final}) ---")
                registrar_log("INFO", "INICIANDO_AUDITORIA", cliente_nome, f"Lote: {lote_final}")
                
                dados_raw_wp = resultados_raw.get(cliente_nome)
                if not dados_raw_wp:
                    registrar_log("ERRO", "FALHA_DADOS_WIDEPAY", cliente_nome, "Dados brutos do WidePay nao localizados.")
                    dados_raw_wp = {
                        "cliente": cliente_nome,
                        "status_conexao": "LOGADO",
                        "carnes": [],
                        "cobrancas": []
                    }
                else:
                    if dados_raw_wp.get("cliente"):
                        dados_contrato["cliente"] = dados_raw_wp["cliente"]
                
                # Normalizar e de-duplicar lançamentos
                valor_base = float(dados_contrato.get("valor_parcela") or 0.0)
                dados_normalizados = normalizar_e_deduplicar(dados_raw_wp, valor_base)
                
                # Reconciliação financeira e cálculos
                dados_calculados = calcular_resumo(dados_normalizados, dados_contrato)
                
                # Validação matemática de regras
                status_val, notas, bloqueado = validar_conciliacao(dados_calculados, dados_contrato)
                
                print(f"Status da Auditoria: {status_val}")
                for nota in notas:
                    print(f"- {nota}")
                    
                # Gerar relatórios finais
                lote_entrega = lote_opcao if (lote_opcao and lote_opcao != "-") else lote_final
                print(f"[DEBUG_SCOPES] lote_opcao={lote_opcao}, lote_final={lote_final}, lote_entrega={lote_entrega}")
                pasta_entrega = preparar_diretorio_entrega(cliente_nome, lote_entrega)
                data_sufixo = datetime.now().strftime("%Y%m%d_%H%M")
                
                dados_contrato_copia = dados_contrato.copy()
                dados_contrato_copia["lote"] = lote_entrega
                
                arquivos = exportar_relatorios_finais(dados_contrato_copia, dados_calculados, dados_normalizados, pasta_entrega, data_sufixo)
                
                registrar_log("SUCESSO", "AUDITORIA_CONCLUIDA", cliente_nome, f"Status: {status_val}; Arquivos gerados na pasta de entrega.")
                
                resultados.append({
                    "cliente": cliente_nome,
                    "lote": lote_entrega,
                    "status": status_val,
                    "bloqueado": bloqueado,
                    "notas": "; ".join(notas),
                    "caminho_entrega": str(pasta_entrega),
                    "arquivos": arquivos
                })
        except Exception as e:
            import traceback
            traceback.print_exc()
            for cli in bloco_info:
                registrar_log("ERRO", "FALHA_AUDITORIA", cli["nome"], str(e))
                print(f"Erro na auditoria do cliente {cli['nome']}: {e}")
                
    if not resultados:
        print("\nERRO: Nenhuma auditoria foi concluida com sucesso.")
        sys.exit(1)
            
    # 3. Gerar consolidação caso solicitado
    if args.consolidado and resultados:
        print("\nGerando relatorio consolidado de todos os clientes processados...")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidado de Auditoria"
        
        headers = ["Cliente", "Lote", "Status Reconciliação", "Bloqueado?", "Notas da Validação", "Diretório de Entrega"]
        ws.append(headers)
        
        # Estilos do header
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for r in resultados:
            ws.append([
                r["cliente"], r["lote"], r["status"],
                "Sim" if r["bloqueado"] else "Não",
                r["notas"], r["caminho_entrega"]
            ])
            
        # Formatar largura das colunas
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        data_sufixo = datetime.now().strftime("%Y%m%d_%H%M")
        caminho_consolidado = ROOT_DIR / "02_RELATORIOS_GERADOS" / f"CONSOLIDADO_AUDITORIA_{data_sufixo}.xlsx"
        wb.save(caminho_consolidado)
        print(f"Relatório consolidado salvo em {caminho_consolidado}")
        
    print("\nAuditoria financeira concluida com sucesso!")

def main():
    try:
        asyncio.run(main_async())
    except KeyboardInterrupt:
        print("\nExecucao interrompida pelo usuario.")
        sys.exit(1)
    except Exception as e:
        print(f"\nErro critico de execucao: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
