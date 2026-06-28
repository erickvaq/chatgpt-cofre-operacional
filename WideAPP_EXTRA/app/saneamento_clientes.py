# -*- coding: utf-8 -*-
"""Camada persistente de saneamento manual dos clientes da WideAPP_EXTRA."""

import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app import config, paths


ARQUIVO_SANEAMENTO = paths.get_internal_data_dir() / "saneamento_clientes.json"
ACOES_MANUAIS = {"quitado", "bloqueado", "ignorar", "atencao"}
ALIAS_NOMES = {
    "emanuel felix da costa filho": "EMMANUEL FELIX DA COSTA FILHO",
    "emmanuel felix da costa filho": "EMMANUEL FELIX DA COSTA FILHO",
}
PLACEHOLDER_RE = re.compile(
    r"(^[\-_.\s]*$|x{3,}\s*y{2,}|^quadra\s+[a-h]$|^total\b|^cliente$|^nome$)",
    re.IGNORECASE,
)


def normalizar(texto):
    texto = str(texto or "")
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", texto).strip()


def slug_busca(texto):
    return re.sub(r"[^a-z0-9]+", " ", normalizar(texto).lower()).strip()


def nome_canonico(nome):
    slug = slug_busca(nome)
    for alias, canonico in ALIAS_NOMES.items():
        if alias in slug:
            return canonico
    return normalizar(nome).upper()


def lote_canonico(registro):
    chave = str(registro.get("chave_lote_canonica") or "").strip().upper()
    if chave and chave not in ("-", "NONE"):
        return chave
    quadra = str(registro.get("quadra") or "").strip().upper()
    lote = str(registro.get("lote") or "").strip().upper()
    if re.fullmatch(r"[A-H]\d{1,3}[A-Z]?", lote):
        return lote
    if quadra and lote and lote not in ("-", "OCORRER", "SER", "DE", "S"):
        return f"{quadra}{lote}".replace(" ", "")
    return lote if lote and lote != "-" else ""


def chave_saneamento(registro):
    nome = nome_canonico(registro.get("cliente") or registro.get("nome") or "")
    lote = lote_canonico(registro)
    return f"{slug_busca(nome)}|{lote}"


def registro_estrutural_ou_placeholder(registro):
    nome = normalizar(registro.get("cliente") or registro.get("nome") or "")
    return bool(PLACEHOLDER_RE.search(nome))


def _linha_para_registro(ws, row):
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    data = {}
    for c, header in enumerate(headers, 1):
        if header:
            data[header] = ws.cell(row, c).value
    return {
        "cliente": data.get("cliente") or data.get("Cliente") or data.get("nome") or "",
        "lote": data.get("lote") or data.get("Lote") or "",
        "quadra": data.get("quadra") or data.get("Quadra") or "",
        "chave_lote_canonica": data.get("chave_lote_canonica") or data.get("chave") or "",
        "contrato_resumo": data.get("contrato_resumo") or "",
        "status": data.get("status") or "",
    }


def _acao_por_cor(ws, row):
    cores = []
    for col in range(1, min(ws.max_column, 12) + 1):
        fg = ws.cell(row, col).fill.fgColor
        if fg.type == "rgb" and fg.rgb:
            cores.append(str(fg.rgb).upper())
        elif fg.type == "indexed":
            try:
                cores.append(f"INDEXED:{int(fg.indexed)}")
            except Exception:
                pass
    if "FF0070C0" in cores:
        return "quitado"
    if "INDEXED:2" in cores:
        return "bloqueado"
    if "INDEXED:5" in cores:
        return "atencao"
    return ""


def encontrar_planilha_marcada():
    candidatos = [
        paths.get_app_root() / "MARCADOS remover BANCO_DADOS_WIDEAPP_EXTRA - Copia.xlsx",
        config.APP_DIR / "MARCADOS remover BANCO_DADOS_WIDEAPP_EXTRA - Copia.xlsx",
        config.APP_DIR.parent / "MARCADOS remover BANCO_DADOS_WIDEAPP_EXTRA - Copia.xlsx",
    ]
    for path in candidatos:
        if path.exists():
            return path
    return None


def carregar_saneamento():
    if not ARQUIVO_SANEAMENTO.exists():
        return {
            "gerado_em": "",
            "decisoes": {},
            "linhas_planilha": {},
            "ordem_planilha": [],
            "movimentos_manuais": {},
            "auditoria_manual": [],
            "resumo": {},
        }
    with open(ARQUIVO_SANEAMENTO, "r", encoding="utf-8") as f:
        payload = json.load(f)
    payload.setdefault("decisoes", {})
    payload.setdefault("linhas_planilha", {})
    payload.setdefault("ordem_planilha", [])
    payload.setdefault("movimentos_manuais", {})
    payload.setdefault("auditoria_manual", [])
    payload.setdefault("resumo", {})
    return payload


def salvar_saneamento(payload):
    ARQUIVO_SANEAMENTO.parent.mkdir(parents=True, exist_ok=True)
    with open(ARQUIVO_SANEAMENTO, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def _configurar_status_operacional(item, acao, origem="", motivo=""):
    acao = acao or "ativo"
    item["saneamento_acao"] = acao
    item["saneamento_origem"] = origem
    item["saneamento_observacao"] = motivo
    item["quitado_manual"] = acao == "quitado"
    item["bloqueado_removido_manual"] = acao == "bloqueado"
    item["atencao_manual"] = acao == "atencao"
    item["ignorado_manual"] = acao == "ignorar"
    if acao == "quitado":
        item["saneamento_categoria"] = "quitados"
        item["ativo_na_lista_principal"] = False
        item["status_operacional"] = "Quitado"
        item["origem_classificacao"] = "manual_interface" if origem == "manual_interface" else origem
        item["bloqueado_reaparecer"] = True
    elif acao == "bloqueado":
        item["saneamento_categoria"] = "bloqueados_removidos"
        item["ativo_na_lista_principal"] = False
        item["status_operacional"] = "Bloqueado"
        item["origem_classificacao"] = "manual_interface" if origem == "manual_interface" else origem
        item["bloqueado_reaparecer"] = True
    elif acao == "ignorar":
        item["saneamento_categoria"] = "ignorados"
        item["ativo_na_lista_principal"] = False
        item["status_operacional"] = "Ignorado"
        item["origem_classificacao"] = origem
        item["bloqueado_reaparecer"] = True
    else:
        item["saneamento_categoria"] = "ativos"
        item["ativo_na_lista_principal"] = True
        item["status_operacional"] = "Ativo"
        item["origem_classificacao"] = "manual_interface" if origem == "manual_interface" else origem
        item["bloqueado_reaparecer"] = False
    if motivo:
        item["motivo_remocao_lista"] = motivo
    return item


def _acao_manual_para_item(item, movimento):
    status = str(movimento.get("status_operacional") or "").strip().lower()
    if status == "quitado":
        return _configurar_status_operacional(
            item,
            "quitado",
            origem="manual_interface",
            motivo=movimento.get("motivo_remocao_lista") or "Movido manualmente para Quitados pelo usuario",
        )
    if status == "bloqueado":
        return _configurar_status_operacional(
            item,
            "bloqueado",
            origem="manual_interface",
            motivo=movimento.get("motivo_remocao_lista") or "Movido manualmente para Bloqueados / Removidos pelo usuario",
        )
    return _configurar_status_operacional(
        item,
        "ativo",
        origem="manual_interface",
        motivo=movimento.get("motivo_remocao_lista") or "Restaurado manualmente para Ativos pelo usuario",
    )


def importar_planilha_marcada():
    payload = carregar_saneamento()
    decisoes = {}
    linhas_planilha = {}
    ordem_planilha = []
    payload["decisoes"] = decisoes
    payload["linhas_planilha"] = linhas_planilha
    payload["ordem_planilha"] = ordem_planilha
    planilha = encontrar_planilha_marcada()
    if not planilha:
        payload["resumo"] = {"planilha_marcada": "nao_encontrada", "decisoes": len(decisoes)}
        salvar_saneamento(payload)
        return payload

    wb = load_workbook(planilha, data_only=True)
    agora = datetime.now().isoformat(timespec="seconds")
    importadas = {"quitado": 0, "bloqueado": 0, "ignorar": 0, "atencao": 0}
    for ws in wb.worksheets:
        for row in range(2, ws.max_row + 1):
            acao = _acao_por_cor(ws, row)
            registro = _linha_para_registro(ws, row)
            nome_visual = str(registro.get("cliente") or "").strip().upper()
            nome = nome_canonico(registro.get("cliente"))
            lote = lote_canonico(registro)
            chave = chave_saneamento(registro)
            if not slug_busca(nome):
                continue
            if acao == "bloqueado" and registro_estrutural_ou_placeholder(registro):
                acao = "ignorar"
            linhas_planilha[chave] = {
                "chave": chave,
                "cliente": nome_visual,
                "lote": str(registro.get("lote") or "").strip().upper(),
                "quadra": str(registro.get("quadra") or "").strip().upper(),
                "lote_canonico": lote,
                "acao": acao,
            }
            ordem_planilha.append(chave)
            if lote:
                linhas_planilha.setdefault(f"lote::{lote}", []).append(chave)
            if not acao:
                continue
            decisoes[chave] = {
                "chave": chave,
                "cliente_canonico": nome_visual,
                "lote_canonico": lote,
                "acao": acao,
                "origem": "planilha_marcada",
                "arquivo_origem": str(planilha),
                "aba_origem": ws.title,
                "linha_origem": row,
                "atualizado_em": agora,
                "observacao": f"Classificacao manual por cor: {acao}",
            }
            importadas[acao] += 1

    payload["gerado_em"] = agora
    payload["resumo"] = {
        "planilha_marcada": str(planilha),
        "decisoes": len(decisoes),
        "importadas": importadas,
    }
    salvar_saneamento(payload)
    return payload


def _decisao_para_registro(registro, decisoes):
    chave = chave_saneamento(registro)
    if chave in decisoes:
        return decisoes[chave]
    return None


def _referencia_planilha(registro, payload):
    linhas = payload.get("linhas_planilha", {})
    chave = chave_saneamento(registro)
    if chave in linhas:
        return linhas[chave]
    lote = lote_canonico(registro)
    candidatos = linhas.get(f"lote::{lote}", [])
    if len(candidatos) == 1:
        return linhas.get(candidatos[0])
    return None


def _acao_automatica(registro):
    if registro_estrutural_ou_placeholder(registro):
        return "ignorar", "placeholder_ou_divisor_estrutural"
    return "", ""


def int_safe(valor):
    try:
        return int(float(str(valor).replace(",", ".")))
    except Exception:
        return 0


def _fabricar_registro_planilha(referencia):
    acao = referencia.get("acao") or "ativo"
    categoria = "ativos"
    if acao == "quitado":
        categoria = "quitados"
    elif acao == "bloqueado":
        categoria = "bloqueados_removidos"
    elif acao == "ignorar":
        categoria = "ignorados"
    return {
        "selecionado": False,
        "cliente": referencia.get("cliente") or "",
        "lote": referencia.get("lote") or "",
        "quadra": referencia.get("quadra") or "",
        "chave_lote_canonica": referencia.get("lote_canonico") or "",
        "status": "",
        "contrato": "Nao encontrado",
        "contrato_arquivo": "",
        "origem": "Planilha marcada",
        "parcelas_pagas_identificadas": "",
        "parcelas_total_contrato": "",
        "parcelas_restantes": "",
        "parcelas_resumo": "",
        "contrato_resumo": "",
        "contrato_modalidade": "",
        "valor_base_parcela": "",
        "valor_total_contratado": "",
        "entrada_valor": "",
        "origem_contrato": "",
        "valor_total_pago": "",
        "ultima_parcela_paga": "",
        "ultimo_vencimento_pago": "",
        "valor_ultimo_pagamento": "",
        "data_atualizacao": "",
        "ultima_atualizacao_widepay": "",
        "situacao_final": "",
        "status_atraso_qtd": "",
        "status_atraso_cor": "",
        "status_atraso_rotulo": "",
        "status_atraso_origem": "",
        "observacoes": "Registro reposto a partir da planilha marcada",
        "divergencias": "",
        "pasta_local": "",
        "boletos_atrasados": "",
        "origem_dados": "Planilha marcada",
        "saneamento_acao": acao,
        "saneamento_categoria": categoria,
        "saneamento_origem": "planilha_marcada",
        "saneamento_observacao": f"Roster restaurado da planilha marcada: {acao}",
        "cliente_canonico": referencia.get("cliente") or "",
        "chave_saneamento": referencia.get("chave") or "",
        "quitado_manual": acao == "quitado",
        "bloqueado_removido_manual": acao == "bloqueado",
        "atencao_manual": acao == "atencao",
        "ignorado_manual": acao == "ignorar",
        "ativo_na_lista_principal": acao in ("", "ativo", "atencao"),
        "status_operacional": "Ativo" if acao in ("", "ativo", "atencao") else "Quitado" if acao == "quitado" else "Bloqueado" if acao == "bloqueado" else "Ignorado",
        "origem_classificacao": "planilha_marcada",
        "bloqueado_reaparecer": acao in ("quitado", "bloqueado", "ignorar"),
        "motivo_remocao_lista": "",
    }


def registrar_movimentacao_manual(registros, chaves, destino, aba_origem):
    payload = importar_planilha_marcada()
    movimentos = payload.setdefault("movimentos_manuais", {})
    auditoria = payload.setdefault("auditoria_manual", [])
    agora = datetime.now().isoformat(timespec="seconds")
    destino_map = {
        "ativos": {
            "status_operacional": "Ativo",
            "ativo_na_lista_principal": True,
            "bloqueado_reaparecer": False,
            "motivo_remocao_lista": "Restaurado manualmente para Ativos pelo usuario",
            "rotulo": "Restaurar para Ativos",
        },
        "quitados": {
            "status_operacional": "Quitado",
            "ativo_na_lista_principal": False,
            "bloqueado_reaparecer": True,
            "motivo_remocao_lista": "Movido manualmente para Quitados pelo usuario",
            "rotulo": "Mover para Quitados",
        },
        "bloqueados_removidos": {
            "status_operacional": "Bloqueado",
            "ativo_na_lista_principal": False,
            "bloqueado_reaparecer": True,
            "motivo_remocao_lista": "Movido manualmente para Bloqueados / Removidos pelo usuario",
            "rotulo": "Mover para Bloqueados / Removidos",
        },
    }
    if destino not in destino_map:
        raise ValueError(f"Destino manual invalido: {destino}")

    por_chave = {}
    for registro in registros:
        item = dict(registro or {})
        ref = _referencia_planilha(item, payload)
        if ref:
            item["cliente"] = ref.get("cliente") or item.get("cliente")
            item["lote"] = ref.get("lote") or item.get("lote")
            item["quadra"] = ref.get("quadra") or item.get("quadra")
            item["chave_lote_canonica"] = ref.get("lote_canonico") or item.get("chave_lote_canonica")
        item["chave_saneamento"] = chave_saneamento(item)
        por_chave[item["chave_saneamento"]] = item

    alterados = []
    for chave in sorted(set(chaves)):
        item = por_chave.get(chave)
        if not item:
            continue
        anterior = item.get("status_operacional") or item.get("saneamento_categoria") or "Ativo"
        movimento = {
            "chave": chave,
            "cliente": item.get("cliente") or "",
            "lote": item.get("lote") or "",
            "quadra": item.get("quadra") or "",
            "chave_lote_canonica": item.get("chave_lote_canonica") or "",
            "aba_origem": aba_origem,
            "aba_destino": destino,
            "status_anterior": anterior,
            "status_operacional": destino_map[destino]["status_operacional"],
            "ativo_na_lista_principal": destino_map[destino]["ativo_na_lista_principal"],
            "origem_classificacao": "manual_interface",
            "bloqueado_reaparecer": destino_map[destino]["bloqueado_reaparecer"],
            "motivo_remocao_lista": destino_map[destino]["motivo_remocao_lista"],
            "atualizado_em": agora,
        }
        movimentos[chave] = movimento
        auditoria.append(
            {
                "data_hora": agora,
                "acao_executada": destino_map[destino]["rotulo"],
                "cliente": item.get("cliente") or "",
                "lote": item.get("lote") or "",
                "quadra": item.get("quadra") or "",
                "chave_lote_canonica": item.get("chave_lote_canonica") or "",
                "aba_origem": aba_origem,
                "aba_destino": destino,
                "status_anterior": anterior,
                "status_novo": destino_map[destino]["status_operacional"],
                "motivo": destino_map[destino]["motivo_remocao_lista"],
                "origem": "manual_interface",
            }
        )
        alterados.append(chave)

    payload["auditoria_manual"] = auditoria[-200:]
    payload["movimentos_manuais"] = movimentos
    salvar_saneamento(payload)
    return alterados


def aplicar_saneamento(registros):
    payload = importar_planilha_marcada()
    decisoes = payload.get("decisoes", {})
    movimentos = payload.get("movimentos_manuais", {})
    saneados = []
    vistos_operacionais = {"quitados": set(), "bloqueados_removidos": set(), "ignorados": set()}

    for registro in registros:
        item = dict(registro or {})
        referencia = _referencia_planilha(item, payload)
        if referencia:
            item["cliente"] = referencia.get("cliente") or item.get("cliente")
            if referencia.get("lote"):
                item["lote"] = referencia.get("lote")
            if referencia.get("quadra"):
                item["quadra"] = referencia.get("quadra")
            if referencia.get("lote_canonico"):
                item["chave_lote_canonica"] = referencia.get("lote_canonico")
        item["cliente_canonico"] = nome_canonico(item.get("cliente"))
        item["chave_saneamento"] = chave_saneamento(item)
        decisao = _decisao_para_registro(item, decisoes)
        acao = ""
        origem = ""
        observacao = ""
        if decisao:
            acao = decisao.get("acao") or ""
            origem = decisao.get("origem") or "planilha_marcada"
            observacao = decisao.get("observacao") or ""
        elif referencia and referencia.get("acao") in ACOES_MANUAIS:
            acao = referencia.get("acao") or ""
            origem = "planilha_marcada"
            observacao = f"Classificacao por lote da planilha marcada: {acao}"
        else:
            acao, origem = _acao_automatica(item)
            observacao = origem

        if acao not in ACOES_MANUAIS:
            acao = "ativo"

        item = _configurar_status_operacional(item, acao, origem=origem, motivo=observacao)
        if origem == "planilha_marcada":
            item["quitado_manual"] = acao == "quitado"
            item["bloqueado_removido_manual"] = acao in ("bloqueado", "ignorar")
            item["atencao_manual"] = acao == "atencao"
            item["ignorado_manual"] = acao == "ignorar"

        movimento_manual = movimentos.get(item.get("chave_saneamento") or "")
        if movimento_manual:
            item = _acao_manual_para_item(item, movimento_manual)

        categoria_item = item.get("saneamento_categoria")
        if categoria_item == "quitados":
            item["saneamento_categoria"] = "quitados"
            chave = item.get("chave_saneamento")
            if chave in vistos_operacionais["quitados"]:
                continue
            vistos_operacionais["quitados"].add(chave)
        elif categoria_item == "bloqueados_removidos":
            item["saneamento_categoria"] = "bloqueados_removidos"
            chave = item.get("chave_saneamento")
            if chave in vistos_operacionais["bloqueados_removidos"]:
                continue
            vistos_operacionais["bloqueados_removidos"].add(chave)
        elif categoria_item == "ignorados":
            item["saneamento_categoria"] = "ignorados"
            chave = item.get("chave_saneamento")
            if chave in vistos_operacionais["ignorados"]:
                continue
            vistos_operacionais["ignorados"].add(chave)
        else:
            item["saneamento_categoria"] = "ativos"
        saneados.append(item)
    linhas_reais = {
        chave: valor
        for chave, valor in payload.get("linhas_planilha", {}).items()
        if not str(chave).startswith("lote::")
    }
    por_chave = {}
    extras = []
    for item in saneados:
        chave = item.get("chave_saneamento") or ""
        if chave in linhas_reais and chave not in por_chave:
            por_chave[chave] = item
        elif chave in linhas_reais:
            extras.append(item)
        else:
            item["saneamento_acao"] = "fora_roster"
            item["saneamento_categoria"] = "fora_roster"
            item["saneamento_origem"] = "planilha_marcada_roster"
            item["saneamento_observacao"] = "Registro fora do roster da planilha marcada"
            extras.append(item)

    resultado = []
    for chave in payload.get("ordem_planilha", []):
        referencia = linhas_reais.get(chave)
        if not referencia:
            continue
        if chave in por_chave:
            resultado.append(por_chave[chave])
        else:
            resultado.append(_fabricar_registro_planilha(referencia))
    resultado.extend(extras)

    resumo = {"ativos": 0, "quitados": 0, "bloqueados_removidos": 0, "ignorados": 0, "fora_roster": 0, "atencao": 0}
    for item in resultado:
        categoria = item.get("saneamento_categoria")
        if categoria == "quitados":
            resumo["quitados"] += 1
        elif categoria == "bloqueados_removidos":
            resumo["bloqueados_removidos"] += 1
        elif categoria == "ignorados":
            resumo["ignorados"] += 1
        elif categoria == "fora_roster":
            resumo["fora_roster"] += 1
        else:
            resumo["ativos"] += 1
            if item.get("saneamento_acao") == "atencao":
                resumo["atencao"] += 1

    payload["ultimo_resumo_aplicado"] = resumo
    salvar_saneamento(payload)
    return resultado, resumo


def separar_por_categoria(registros):
    grupos = {"ativos": [], "quitados": [], "bloqueados_removidos": [], "ignorados": [], "fora_roster": [], "atencao": []}
    for item in registros:
        categoria = item.get("saneamento_categoria") or "ativos"
        if categoria not in grupos:
            categoria = "ativos"
        grupos[categoria].append(item)
        if item.get("saneamento_acao") == "atencao":
            grupos["atencao"].append(item)
    return grupos
