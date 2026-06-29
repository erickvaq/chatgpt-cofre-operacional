# -*- coding: utf-8 -*-
"""Indexacao local de clientes/lotes para a interface dinamica."""

import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app import config, saneamento_clientes
from app.leitor_contratos import (
    extrair_texto_docx,
    extrair_texto_pdf,
    extrair_nome_cliente,
    parsear_dados_contrato,
)
from app.normalizador_pagamentos import normalizar_e_deduplicar
from app.widepay_boletos_cache import (
    chave_lote_canonica,
    montar_raw_cliente,
    normalizar_lote_quadra,
    registros_de_resultado_bloco,
    salvar_cache as salvar_cache_boletos_widepay,
)


LOTE_RE = re.compile(r"\b([A-H]\s*[-.]?\s*\d{1,3}[A-Z]?)\b", re.IGNORECASE)
QUADRA_RE = re.compile(r"\bquadra\s*([A-H])\b", re.IGNORECASE)
CONTRATO_A_VISTA_RE = re.compile(
    r"\b(a\s+vista|avista|pagamento\s+a\s+vista|pagamento\s+avista|quitad[oa]?|quitacao|integralmente\s+pago|valor\s+integral|sem\s+parcelamento|parcela\s+unica)\b",
    re.IGNORECASE,
)
CONTRATO_QUITADO_RE = re.compile(r"\b(quitad[oa]?|quitacao|sem parcelas pendentes)\b", re.IGNORECASE)
PASTA_IGNORADA_RE = re.compile(r"\b(backup|bkp|nova pasta|temporari[oa])\b", re.IGNORECASE)
COMPENSACAO_ATRASO_RE = re.compile(r"\b(atras[oa]s?|atrazos?|ref(?:erente)?|referencia)\b", re.IGNORECASE)
REMOVE_WORDS_RE = re.compile(
    r"\b(contrato|copia|final|corrigido|previa|carne\d*|apart\d*|agua|viva|leandro|meirelles|leo|docx|pdf|txt|html|md)\b",
    re.IGNORECASE,
)
MESES_ABREV = ("Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez")
COLUNAS_XLSX_PAGAMENTOS_RECENTES = [
    ("cliente", "Cliente"),
    ("lote", "Lote / Quadra"),
    ("contrato_resumo", "Contrato"),
    ("parcelas_resumo", "Parcelas"),
    ("situacao_final", "Situacao"),
    ("ultima_atualizacao_widepay", "Atualizado em"),
    ("valor_total_contratado", "Valor Lote"),
    ("valor_total_pago", "Total pago"),
]

PERSISTIR_CAMPOS = (
    "status",
    "parcelas_pagas_identificadas",
    "parcelas_total_contrato",
    "parcelas_restantes",
    "parcelas_resumo",
    "contrato_resumo",
    "contrato_modalidade",
    "valor_base_parcela",
    "valor_total_contratado",
    "entrada_valor",
    "origem_contrato",
    "valor_total_pago",
    "ultima_parcela_paga",
    "ultimo_vencimento_pago",
    "valor_ultimo_pagamento",
    "ultima_atualizacao_widepay",
    "situacao_final",
    "status_atraso_qtd",
    "status_atraso_cor",
    "status_atraso_rotulo",
    "status_atraso_origem",
    "observacoes",
    "divergencias",
    "boletos_atrasados",
    "origem_dados",
    "saneamento_acao",
    "saneamento_categoria",
    "saneamento_origem",
    "saneamento_observacao",
    "cliente_canonico",
    "chave_saneamento",
    "quitado_manual",
    "bloqueado_removido_manual",
    "atencao_manual",
    "ignorado_manual",
    "ativo_na_lista_principal",
    "status_operacional",
    "origem_classificacao",
    "bloqueado_reaparecer",
    "motivo_remocao_lista",
    "pagamentos_recentes_5m",
)

PRESERVAR_EM_REINDEXACAO = (
    "selecionado",
    "cliente",
    "status",
    "parcelas_pagas_identificadas",
    "parcelas_total_contrato",
    "parcelas_restantes",
    "parcelas_resumo",
    "contrato_resumo",
    "contrato_modalidade",
    "valor_base_parcela",
    "valor_total_contratado",
    "entrada_valor",
    "origem_contrato",
    "valor_total_pago",
    "ultima_parcela_paga",
    "ultimo_vencimento_pago",
    "valor_ultimo_pagamento",
    "data_atualizacao",
    "ultima_atualizacao_widepay",
    "situacao_final",
    "status_atraso_qtd",
    "status_atraso_cor",
    "status_atraso_rotulo",
    "status_atraso_origem",
    "observacoes",
    "divergencias",
    "boletos_atrasados",
    "origem_dados",
    "ativo_na_lista_principal",
    "status_operacional",
    "origem_classificacao",
    "bloqueado_reaparecer",
    "motivo_remocao_lista",
    "pagamentos_recentes_5m",
)


CAMPOS_NUNCA_ESVAZIAR = set(PERSISTIR_CAMPOS) | {
    "cliente",
    "lote",
    "quadra",
    "contrato",
    "contrato_arquivo",
    "pasta_local",
}


def normalizar(texto):
    texto = texto or ""
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", texto).strip()


def slug_busca(texto):
    return re.sub(r"[^a-z0-9]+", " ", normalizar(texto).lower()).strip()


def inteiro(valor, default=0):
    if valor in (None, ""):
        return default
    try:
        return int(float(str(valor).replace(",", ".")))
    except (TypeError, ValueError):
        return default


def decimal(valor, default=0.0):
    if valor in (None, ""):
        return default
    try:
        if isinstance(valor, str):
            limpo = valor.replace("R$", "").replace(".", "").replace(",", ".").strip()
            return float(limpo)
        return float(valor)
    except (TypeError, ValueError):
        return default


def valor_preenchido(valor):
    if valor is None:
        return False
    if isinstance(valor, str):
        return valor.strip() not in ("", "-", "None", "NAO CONFIRMADO", "Nao encontrado")
    return True


def aplicar_se_preenchido(registro, campo, valor):
    if valor_preenchido(valor):
        registro[campo] = valor


def formatar_moeda(valor):
    if valor in (None, ""):
        return ""
    numero = decimal(valor, None)
    if numero is None:
        return str(valor)
    return f"R$ {numero:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def formatar_data_hora(valor):
    if not valor:
        return "Nunca atualizado"
    bruto = str(valor).strip()
    for parser in (datetime.fromisoformat,):
        try:
            return parser(bruto).strftime("%d/%m/%Y %H:%M")
        except ValueError:
            pass
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(bruto, fmt).strftime("%d/%m/%Y %H:%M")
        except ValueError:
            pass
    return bruto


def parse_data(valor):
    if not valor:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    bruto = str(valor).strip()
    if not bruto or bruto in ("-", "None"):
        return None
    try:
        return datetime.fromisoformat(bruto.replace("Z", "+00:00")).date()
    except ValueError:
        pass
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(bruto, fmt).date()
        except ValueError:
            pass
    return None


def _deslocar_mes(ano, mes, deslocamento):
    indice = (ano * 12 + (mes - 1)) + deslocamento
    novo_ano = indice // 12
    novo_mes = (indice % 12) + 1
    return novo_ano, novo_mes


def janela_pagamentos_recentes(quantidade=10, referencia=None):
    referencia = referencia or datetime.now().date()
    meses = []
    inicio = -(max(1, quantidade) - 1)
    for deslocamento in range(inicio, 1):
        ano, mes = _deslocar_mes(referencia.year, referencia.month, deslocamento)
        chave = f"{ano:04d}-{mes:02d}"
        rotulo = f"{MESES_ABREV[mes - 1]}/{str(ano)[-2:]}"
        meses.append({"chave": chave, "rotulo": rotulo, "ano": ano, "mes": mes})
    return meses


def obter_celula_pagamento_recente(item, mes):
    mapa = item.get("pagamentos_recentes_10m") or item.get("pagamentos_recentes_5m") or {}
    if not isinstance(mapa, dict):
        mapa = {}
        
    chave = mes["chave"]
    mes_data = mapa.get(chave)
    cliente_nome = str(item.get("cliente", "")).strip()
    valor_base = decimal(item.get("valor_base_parcela"))
    
    # 1. Se for dicionario rico
    if isinstance(mes_data, dict) and "status" in mes_data:
        status_str = mes_data.get("status", "Sem boleto")
        texto1 = mes_data.get("texto1", "-")
        texto2 = mes_data.get("texto2", "-")
        
        # Se for Pago mas sem valor real, OU se for "Sem boleto" (para curar cache corrompido),
        # tentamos buscar no raw/WidePay
        if (status_str in ("Pago", "Recebido") and (texto1 in ("Pago", "Recebido", "-", "Sem boleto"))) or (status_str == "Sem boleto"):
            if cliente_nome:
                raw = montar_raw_cliente(
                    cliente=cliente_nome,
                    lote=item.get("lote", ""),
                    quadra=item.get("quadra", ""),
                    pasta_local=item.get("pasta_local", ""),
                )
                if raw:
                    normalizado = normalizar_e_deduplicar(raw, valor_base=valor_base)
                    cobrancas = normalizado.get("cobrancas") or []
                    encontrou = False
                    for cob in cobrancas:
                        if cob.get("duplicidade"):
                            continue
                        dt_ref = parse_data(cob.get("vencimento")) or parse_data(cob.get("pagamento"))
                        if dt_ref and f"{dt_ref.year:04d}-{dt_ref.month:02d}" == chave:
                            bruto = normalizar(str(cob.get("status") or "")).lower()
                            if bruto in ("recebido", "pago", "quitado", "liquidado"):
                                v_pago = cob.get("valor_recebido") or cob.get("recebido") or cob.get("valor_pago") or cob.get("valor") or valor_base
                                texto1 = formatar_moeda(v_pago)
                                status_str = "Pago"
                            elif bruto in ("vencido",):
                                texto1 = "Vencido"
                                status_str = "Vencido"
                            elif bruto in ("aguardando", "aberto", "pendente", "pendencia"):
                                ref_date = dt_ref.date() if hasattr(dt_ref, "date") else dt_ref
                                if ref_date < datetime.now().date():
                                    texto1 = "Vencido"
                                    status_str = "Vencido"
                                else:
                                    texto1 = "Pendente"
                                    status_str = "Pendente"
                            texto2 = f"{dt_ref.day:02d}/{dt_ref.month:02d}"
                            encontrou = True
                            break
                    if not encontrou and status_str == "Sem boleto":
                        texto1 = "Sem boleto"
                        texto2 = "-"
            if status_str in ("Pago", "Recebido") and (texto1 in ("Pago", "Recebido", "-", "Sem boleto")) and valor_base:
                texto1 = formatar_moeda(valor_base)
                
        return {"status": status_str, "texto1": texto1, "texto2": texto2}
        
    # 2. Se for string (formato legado)
    if mes_data is not None:
        v_str = str(mes_data).strip()
        
        if v_str in ("Pago", "Recebido", "Recebida"):
            val_texto = "-"
            dia_texto = "-"
            if cliente_nome:
                raw = montar_raw_cliente(
                    cliente=cliente_nome,
                    lote=item.get("lote", ""),
                    quadra=item.get("quadra", ""),
                    pasta_local=item.get("pasta_local", ""),
                )
                if raw:
                    normalizado = normalizar_e_deduplicar(raw, valor_base=valor_base)
                    cobrancas = normalizado.get("cobrancas") or []
                    for cob in cobrancas:
                        if cob.get("duplicidade"):
                            continue
                        dt_ref = parse_data(cob.get("vencimento")) or parse_data(cob.get("pagamento"))
                        if dt_ref and f"{dt_ref.year:04d}-{dt_ref.month:02d}" == chave:
                            v_pago = cob.get("valor_recebido") or cob.get("recebido") or cob.get("valor_pago") or cob.get("valor")
                            val_texto = formatar_moeda(v_pago)
                            dia_texto = f"{dt_ref.day:02d}/{dt_ref.month:02d}"
                            break
            if val_texto == "-" and valor_base:
                val_texto = formatar_moeda(valor_base)
            if val_texto == "-":
                val_texto = "Pago"
            return {"status": "Pago", "texto1": val_texto, "texto2": dia_texto}
            
        elif v_str == "Vencido":
            dia_texto = "-"
            if cliente_nome:
                raw = montar_raw_cliente(
                    cliente=cliente_nome,
                    lote=item.get("lote", ""),
                    quadra=item.get("quadra", ""),
                    pasta_local=item.get("pasta_local", ""),
                )
                if raw:
                    normalizado = normalizar_e_deduplicar(raw, valor_base=valor_base)
                    cobrancas = normalizado.get("cobrancas") or []
                    for cob in cobrancas:
                        if cob.get("duplicidade"):
                            continue
                        dt_ref = parse_data(cob.get("vencimento")) or parse_data(cob.get("pagamento"))
                        if dt_ref and f"{dt_ref.year:04d}-{dt_ref.month:02d}" == chave:
                            dia_texto = f"{dt_ref.day:02d}/{dt_ref.month:02d}"
                            break
            return {"status": "Vencido", "texto1": "Vencido", "texto2": dia_texto}
            
        elif v_str == "Pendente":
            dia_texto = "-"
            if cliente_nome:
                raw = montar_raw_cliente(
                    cliente=cliente_nome,
                    lote=item.get("lote", ""),
                    quadra=item.get("quadra", ""),
                    pasta_local=item.get("pasta_local", ""),
                )
                if raw:
                    normalizado = normalizar_e_deduplicar(raw, valor_base=valor_base)
                    cobrancas = normalizado.get("cobrancas") or []
                    for cob in cobrancas:
                        if cob.get("duplicidade"):
                            continue
                        dt_ref = parse_data(cob.get("vencimento")) or parse_data(cob.get("pagamento"))
                        if dt_ref and f"{dt_ref.year:04d}-{dt_ref.month:02d}" == chave:
                            dia_texto = f"{dt_ref.day:02d}/{dt_ref.month:02d}"
                            break
            return {"status": "Pendente", "texto1": "Pendente", "texto2": dia_texto}
            
        elif v_str in ("Sem boleto", "-", ""):
            return {"status": "Sem boleto", "texto1": "Sem boleto", "texto2": "-"}
            
    # 3. Fallback: procurar no WidePay brutos se o cliente for informado
    if cliente_nome:
        raw = montar_raw_cliente(
            cliente=cliente_nome,
            lote=item.get("lote", ""),
            quadra=item.get("quadra", ""),
            pasta_local=item.get("pasta_local", ""),
        )
        if raw:
            normalizado = normalizar_e_deduplicar(raw, valor_base=valor_base)
            cobrancas = normalizado.get("cobrancas") or []
            for cob in cobrancas:
                if cob.get("duplicidade"):
                    continue
                dt_ref = parse_data(cob.get("vencimento")) or parse_data(cob.get("pagamento"))
                if dt_ref and f"{dt_ref.year:04d}-{dt_ref.month:02d}" == chave:
                    bruto = normalizar(str(cob.get("status") or "")).lower()
                    if bruto in ("recebido", "pago", "quitado", "liquidado"):
                        v_pago = cob.get("valor_recebido") or cob.get("recebido") or cob.get("valor_pago") or cob.get("valor") or valor_base
                        val_texto = formatar_moeda(v_pago)
                        return {"status": "Pago", "texto1": val_texto, "texto2": f"{dt_ref.day:02d}/{dt_ref.month:02d}"}
                    elif bruto in ("vencido",):
                        return {"status": "Vencido", "texto1": "Vencido", "texto2": f"{dt_ref.day:02d}/{dt_ref.month:02d}"}
                    elif bruto in ("aguardando", "aberto", "pendente", "pendencia"):
                        ref_date = dt_ref.date() if hasattr(dt_ref, "date") else dt_ref
                        if ref_date < datetime.now().date():
                            return {"status": "Vencido", "texto1": "Vencido", "texto2": f"{dt_ref.day:02d}/{dt_ref.month:02d}"}
                        else:
                            return {"status": "Pendente", "texto1": "Pendente", "texto2": f"{dt_ref.day:02d}/{dt_ref.month:02d}"}

    return {"status": "Sem boleto", "texto1": "Sem boleto", "texto2": "-"}


def normalizar_pagamentos_recentes_5m(mapa=None, quantidade=10):
    atual = mapa if isinstance(mapa, dict) else {}
    normalizado = {}
    dummy_item = {"pagamentos_recentes_5m": atual}
    for mes in janela_pagamentos_recentes(quantidade=quantidade):
        normalizado[mes["chave"]] = obter_celula_pagamento_recente(dummy_item, mes)
    if isinstance(mapa, dict) and mapa.get("_enriquecido"):
        normalizado["_enriquecido"] = True
    return normalizado


def meses_desde_data(valor):
    data = parse_data(valor)
    if not data:
        return None
    hoje = datetime.now().date()
    return (hoje.year - data.year) * 12 + (hoje.month - data.month) - (1 if hoje.day < data.day else 0)


def sem_pagamento_recente_parcelado(registro, meses_limite=8):
    modalidade = registro.get("contrato_modalidade") or inferir_modalidade_contrato(
        registro.get("contrato_arquivo"),
        registro.get("pasta_local"),
    )
    if modalidade != "parcelado":
        return False
    total = inteiro(registro.get("parcelas_total_contrato"))
    pagas = inteiro(registro.get("parcelas_pagas_identificadas"))
    meses_sem_pagamento = meses_desde_data(registro.get("ultimo_vencimento_pago"))
    return (
        meses_sem_pagamento is not None
        and meses_sem_pagamento >= meses_limite
        and total > 0
        and pagas < total
    )


def _status_pagamento_recente(status):
    bruto = normalizar(str(status or "")).lower()
    if bruto in ("recebido", "pago", "quitado", "liquidado"):
        return "Pago"
    if bruto in ("vencido", "aguardando", "aberto", "pendente", "pendencia"):
        return "Vencido"
    return ""




def calcular_pagamentos_recentes_5m(cobrancas, valor_base=None, quantidade=10):
    resultado = {mes["chave"]: {"status": "Sem boleto", "texto1": "Sem boleto", "texto2": "-"} for mes in janela_pagamentos_recentes(quantidade=quantidade)}
    prioridade = {"Sem boleto": 0, "Pendente": 1, "Vencido": 2, "Pago": 3}

    for cob in cobrancas or []:
        if cob.get("duplicidade"):
            continue
        classificacao = normalizar(str(cob.get("classificacao") or "")).lower()
        if classificacao in ("entrada", "atraso", "avulso"):
            continue
        if classificacao not in ("parcela normal", "pendencia") and not cob.get("pertence_a_carne"):
            continue

        data_ref = parse_data(cob.get("vencimento")) or parse_data(cob.get("pagamento"))
        if not data_ref:
            continue

        bruto = normalizar(str(cob.get("status") or "")).lower()
        if bruto in ("recebido", "pago", "quitado", "liquidado"):
            status_str = "Pago"
        elif bruto in ("vencido",):
            status_str = "Vencido"
        elif bruto in ("aguardando", "aberto", "pendente", "pendencia"):
            ref_date = data_ref.date() if hasattr(data_ref, "date") else data_ref
            if ref_date < datetime.now().date():
                status_str = "Vencido"
            else:
                status_str = "Pendente"
        else:
            continue

        chave = f"{data_ref.year:04d}-{data_ref.month:02d}"
        if chave not in resultado:
            continue
        
        atual_status = resultado[chave]["status"]
        if prioridade[status_str] >= prioridade[atual_status]:
            if status_str == "Pago":
                v_pago = cob.get("valor_recebido") or cob.get("recebido") or cob.get("valor_pago") or cob.get("valor") or valor_base
                texto1 = formatar_moeda(v_pago)
                if texto1 == "-":
                    texto1 = "Pago"
            elif status_str == "Pendente":
                texto1 = "Pendente"
            elif status_str == "Vencido":
                texto1 = "Vencido"
            else:
                texto1 = "Sem boleto"
                
            texto2 = f"{data_ref.day:02d}/{data_ref.month:02d}"
            resultado[chave] = {"status": status_str, "texto1": texto1, "texto2": texto2}

    res = normalizar_pagamentos_recentes_5m(resultado, quantidade=quantidade)
    res["_enriquecido"] = True
    return res


def enriquecer_pagamentos_recentes_cache(registro):
    item = normalizar_registro_cache(registro)
    atual = item.get("pagamentos_recentes_5m") if isinstance(item.get("pagamentos_recentes_5m"), dict) else {}
    raw = montar_raw_cliente(
        cliente=item.get("cliente", ""),
        lote=item.get("lote", ""),
        quadra=item.get("quadra", ""),
        pasta_local=item.get("pasta_local", ""),
    )
    if not raw:
        res = normalizar_pagamentos_recentes_5m(atual, quantidade=10)
        res["_enriquecido"] = True
        item["pagamentos_recentes_5m"] = res
        return item

    valor_base = decimal(item.get("valor_base_parcela"))
    normalizado = normalizar_e_deduplicar(raw, valor_base=valor_base)
    item["pagamentos_recentes_5m"] = calcular_pagamentos_recentes_5m(normalizado.get("cobrancas") or [], valor_base=valor_base, quantidade=10)
    return item


def classificar_status_atraso(qtd):
    qtd = inteiro(qtd)
    if qtd <= 3:
        return "verde"
    if qtd <= 5:
        return "amarelo"
    return "vermelho"


def rotulo_status_atraso(qtd):
    qtd = inteiro(qtd)
    cor = classificar_status_atraso(qtd)
    if cor == "verde":
        return "Em dia" if qtd == 0 else "Atraso leve"
    if cor == "amarelo":
        return "Atraso"
    return "Atraso critico"


def formatar_status_indicador(registro):
    valor = registro.get("status_atraso_qtd", registro.get("boletos_atrasados", 0))
    return str(inteiro(valor))


def calcular_atraso_recente(metrics, dias_recentes=210):
    normalizado = (metrics or {}).get("normalizado") or {}
    calculos = (metrics or {}).get("calculos") or {}
    cobrancas = normalizado.get("cobrancas") or []
    hoje = datetime.now().date()
    vencidos_recentes = 0

    for cob in cobrancas:
        status = normalizar(str(cob.get("status", ""))).lower()
        vencimento = parse_data(cob.get("vencimento") or cob.get("data_vencimento"))
        if not vencimento or vencimento >= hoje:
            continue
        if (hoje - vencimento).days > dias_recentes:
            continue
        if status in ("recebido", "pago", "cancelado", "baixado"):
            continue
        vencidos_recentes += 1

    compensacoes = 0
    pagamentos = list(calculos.get("pagamentos_interpretados") or [])
    pagamentos.extend(cob for cob in cobrancas if normalizar(str(cob.get("status", ""))).lower() in ("recebido", "pago"))

    for pgto in pagamentos:
        texto = " ".join(
            str(pgto.get(campo, ""))
            for campo in ("descricao", "referencia", "observacao", "cliente", "forma")
        )
        if not COMPENSACAO_ATRASO_RE.search(normalizar(texto).lower()):
            continue
        data_ref = parse_data(pgto.get("pagamento") or pgto.get("data_pagamento") or pgto.get("vencimento"))
        if data_ref and (hoje - data_ref).days > dias_recentes:
            continue
        if decimal(pgto.get("valor_recebido", pgto.get("recebido", pgto.get("valor")))) <= 0:
            continue
        compensacoes += max(1, inteiro(pgto.get("qtd"), 1))

    return max(0, vencidos_recentes - compensacoes)


def extrair_lote(nome):
    match = LOTE_RE.search(nome or "")
    if not match:
        return "-"
    return re.sub(r"\s+", "", match.group(1).upper().replace(".", "").replace("-", ""))


def extrair_quadra(nome, lote):
    match = QUADRA_RE.search(nome or "")
    if match:
        return match.group(1).upper()
    if lote and lote != "-":
        return lote[0].upper()
    return "-"


def limpar_nome_cliente(nome):
    sem_ext = re.sub(r"\.[a-z0-9]{2,5}$", "", nome or "", flags=re.IGNORECASE)
    sem_lote = LOTE_RE.sub(" ", sem_ext)
    sem_ruido = REMOVE_WORDS_RE.sub(" ", sem_lote)
    sem_ruido = sem_ruido.replace("_", " ").replace("-", " ")
    sem_ruido = re.sub(r"\s+", " ", sem_ruido).strip(" ._-")
    return normalizar(sem_ruido).title() if sem_ruido else normalizar(nome).title()


def detectar_contrato(pasta):
    arquivos = []
    for path in Path(pasta).glob("*"):
        if not path.is_file():
            continue
        nome = path.name.lower()
        if path.suffix.lower() not in (".docx", ".pdf", ".txt"):
            continue
        if any(skip in nome for skip in ("carne", "recibo", "wp-pdf", "boleto")):
            continue
        arquivos.append(path)
    if not arquivos:
        return "Nao encontrado", ""
    
    # Heurística de pontuação baseada em correspondência com o nome da pasta
    folder_name = pasta.name.lower()
    noise = {"agua", "viva", "leandro", "meirelles", "leo", "contrato", "lote", "quadra", "lt", "qd"}
    lote_da_pasta = extrair_lote(pasta.name)
    if lote_da_pasta != "-":
        noise.add(lote_da_pasta.lower())
    
    tokens = [t for t in re.split(r"[^a-z0-9]+", folder_name) if t and t not in noise]
    
    def score_candidato(path):
        f_name = path.name.lower()
        score = sum(1 for t in tokens if t in f_name)
        # Prioridade de extensão como peso fracionário
        ext = path.suffix.lower()
        ext_weight = 0.9 if ext == ".docx" else (0.5 if ext == ".pdf" else 0.1)
        # Tamanho do arquivo como peso menor (desempate)
        size_weight = (path.stat().st_size if path.exists() else 0) / 1000000000.0
        return (score, ext_weight, size_weight)
        
    principal = sorted(arquivos, key=score_candidato, reverse=True)[0]
    return "Encontrado", str(principal)


def inferir_modalidade_contrato(contrato_path="", pasta_local="", dados_contrato=None):
    dados_contrato = dados_contrato or {}
    total_parcelas = inteiro(dados_contrato.get("total_parcelas"))
    if total_parcelas == 1:
        return "a_vista"
    referencia = normalizar(" ".join(filter(None, [str(contrato_path), str(pasta_local)]))).lower()
    if CONTRATO_A_VISTA_RE.search(referencia):
        return "a_vista"
    if CONTRATO_QUITADO_RE.search(referencia):
        return "quitado"
    if contrato_path:
        return "parcelado"
    return "nao_confirmado"


def deduzir_situacao_final(registro):
    contrato_status = registro.get("contrato")
    if contrato_status != "Encontrado":
        return "Sem contrato confirmado"

    status = str(registro.get("status") or "").strip().upper()
    total = inteiro(registro.get("parcelas_total_contrato"))
    pagas = inteiro(registro.get("parcelas_pagas_identificadas"))
    atualizado = bool(registro.get("ultima_atualizacao_widepay"))

    if status == "ERRO":
        return "Bloqueado"
    if sem_pagamento_recente_parcelado(registro):
        return "Bloqueado"
    if status == "PENDENTE":
        if total > 0 and pagas >= total:
            return "Quitado com alerta"
        if atualizado and total > 0:
            return "Em andamento com alerta"
        return "Pendente de validacao"
    if total > 0 and pagas >= total:
        return "Quitado"
    if atualizado and total > 0:
        return "Em andamento"
    if atualizado:
        return "Atualizado"
    return "Pendente de auditoria"


def deduzir_resumo_contrato(registro):
    contrato_status = registro.get("contrato")
    if contrato_status != "Encontrado":
        return "NAO CONFIRMADO"

    modalidade = registro.get("contrato_modalidade") or inferir_modalidade_contrato(
        registro.get("contrato_arquivo"),
        registro.get("pasta_local"),
    )
    total = inteiro(registro.get("parcelas_total_contrato"))
    pagas = inteiro(registro.get("parcelas_pagas_identificadas"))
    restantes = inteiro(registro.get("parcelas_restantes"))

    if modalidade == "a_vista":
        return "A VISTA"
    if modalidade == "quitado":
        return "QUITADO"
    if total > 0 and pagas >= total:
        return "CONTRATO QUITADO"
    if total > 0 and pagas > 0 and restantes == 0:
        return "SEM PARCELAS PENDENTES"
    if total > 0 or modalidade == "parcelado":
        return "Parcelado"
    return "NAO CONFIRMADO"


def deduzir_resumo_parcelas(registro):
    contrato_status = registro.get("contrato")
    if contrato_status != "Encontrado":
        return "NAO CONFIRMADO"

    modalidade = registro.get("contrato_modalidade") or inferir_modalidade_contrato(
        registro.get("contrato_arquivo"),
        registro.get("pasta_local"),
    )
    total = inteiro(registro.get("parcelas_total_contrato"))
    pagas = inteiro(registro.get("parcelas_pagas_identificadas"))

    if modalidade == "a_vista":
        return "A VISTA / quitado"
    if modalidade == "quitado":
        return "QUITADO"
    if total <= 0:
        if modalidade == "parcelado":
            return f"{pagas} / ? pagas"
        return "NAO CONFIRMADO"
    if pagas >= total:
        return f"{pagas} / {total} quitado"
    return f"{pagas} / {total} pagas"


def normalizar_registro_cache(registro):
    item = dict(registro or {})
    item.setdefault("selecionado", False)
    item.setdefault("cliente", "")
    item.setdefault("lote", "-")
    item.setdefault("quadra", "-")
    item.setdefault("chave_lote_canonica", "")
    item.setdefault("status", "Pendente validacao WidePay" if item.get("contrato") == "Encontrado" else "Sem contrato confirmado")
    item.setdefault("contrato", "Nao encontrado")
    item.setdefault("contrato_arquivo", "")
    item.setdefault("origem", "Contrato local")
    item.setdefault("parcelas_pagas_identificadas", "")
    item.setdefault("parcelas_total_contrato", "")
    item.setdefault("parcelas_restantes", "")
    item.setdefault("parcelas_resumo", "")
    item.setdefault("contrato_resumo", "")
    item.setdefault("contrato_modalidade", "")
    item.setdefault("valor_base_parcela", "")
    item.setdefault("valor_total_contratado", "")
    item.setdefault("entrada_valor", "")
    item.setdefault("origem_contrato", "")
    item.setdefault("valor_total_pago", "")
    item.setdefault("ultima_parcela_paga", "")
    item.setdefault("ultimo_vencimento_pago", "")
    item.setdefault("valor_ultimo_pagamento", "")
    item.setdefault("data_atualizacao", "")
    item.setdefault("ultima_atualizacao_widepay", "")
    item.setdefault("situacao_final", "")
    item.setdefault("status_atraso_qtd", item.get("boletos_atrasados", 0))
    item.setdefault("status_atraso_cor", "")
    item.setdefault("status_atraso_rotulo", "")
    item.setdefault("status_atraso_origem", "")
    item.setdefault("observacoes", "")
    item.setdefault("divergencias", "")
    item.setdefault("pasta_local", "")
    item.setdefault("boletos_atrasados", "")
    item.setdefault("origem_dados", "")
    item.setdefault("saneamento_acao", "ativo")
    item.setdefault("saneamento_categoria", "ativos")
    item.setdefault("saneamento_origem", "")
    item.setdefault("saneamento_observacao", "")
    item.setdefault("cliente_canonico", "")
    item.setdefault("chave_saneamento", "")
    item.setdefault("quitado_manual", False)
    item.setdefault("bloqueado_removido_manual", False)
    item.setdefault("atencao_manual", False)
    item.setdefault("ignorado_manual", False)
    item.setdefault("ativo_na_lista_principal", True)
    item.setdefault("status_operacional", "Ativo")
    item.setdefault("origem_classificacao", "")
    item.setdefault("bloqueado_reaparecer", False)
    item.setdefault("motivo_remocao_lista", "")
    item["pagamentos_recentes_5m"] = normalizar_pagamentos_recentes_5m(item.get("pagamentos_recentes_5m"))

    if not item.get("contrato_modalidade"):
        item["contrato_modalidade"] = inferir_modalidade_contrato(
            item.get("contrato_arquivo"),
            item.get("pasta_local"),
        )
    if item.get("contrato_modalidade") == "a_vista":
        item["contrato_resumo"] = "A VISTA"
        item["parcelas_resumo"] = "A VISTA / quitado"
        item["parcelas_restantes"] = 0
        item["parcelas_total_contrato"] = item.get("parcelas_total_contrato") or 1
    else:
        total_contrato = inteiro(item.get("parcelas_total_contrato"))
        pagas_contrato = inteiro(item.get("parcelas_pagas_identificadas"))
        if total_contrato > 0 and pagas_contrato >= 0:
            item["parcelas_restantes"] = max(0, total_contrato - pagas_contrato)
    item["chave_lote_canonica"] = normalizar_lote_quadra(
        item.get("lote"),
        item.get("quadra"),
        item.get("referencia"),
        item.get("pasta_local"),
    ) or item.get("chave_lote_canonica") or ""
    situacao_calculada = deduzir_situacao_final(item)
    if situacao_calculada == "Bloqueado" or not item.get("situacao_final"):
        item["situacao_final"] = situacao_calculada
    item["contrato_resumo"] = deduzir_resumo_contrato(item)
    item["parcelas_resumo"] = deduzir_resumo_parcelas(item)
    if sem_pagamento_recente_parcelado(item):
        meses_sem_pagamento = meses_desde_data(item.get("ultimo_vencimento_pago"))
        aviso = f"Sem pagamento recebido ha mais de 8 meses ({meses_sem_pagamento} meses)"
        obs = item.get("observacoes") or ""
        if aviso not in obs:
            item["observacoes"] = f"{obs} | {aviso}" if obs else aviso
        item["status_atraso_cor"] = "vermelho"
        item["status_atraso_rotulo"] = "Sem pagamento recente"
        item["status_atraso_origem"] = "WidePay/cache"
        item["status_atraso_qtd"] = max(inteiro(item.get("status_atraso_qtd")), 6)
    elif item.get("contrato_modalidade") == "parcelado" and item.get("situacao_final") == "Bloqueado":
        item["status_atraso_cor"] = "vermelho"
        item["status_atraso_rotulo"] = item.get("status_atraso_rotulo") or "Bloqueado"
        item["status_atraso_origem"] = item.get("status_atraso_origem") or "WidePay/cache"
        item["status_atraso_qtd"] = max(inteiro(item.get("status_atraso_qtd")), 6)
    item["status_atraso_qtd"] = inteiro(item.get("status_atraso_qtd", item.get("boletos_atrasados", 0)))
    item["boletos_atrasados"] = item["status_atraso_qtd"]
    item["status_atraso_cor"] = item.get("status_atraso_cor") or classificar_status_atraso(item["status_atraso_qtd"])
    item["status_atraso_rotulo"] = item.get("status_atraso_rotulo") or rotulo_status_atraso(item["status_atraso_qtd"])
    item["status_atraso_origem"] = item.get("status_atraso_origem") or ("WidePay/cache" if item.get("ultima_atualizacao_widepay") else "Pendente WidePay")
    if not item.get("origem_dados"):
        item["origem_dados"] = "Contrato local + cache" if item.get("ultima_atualizacao_widepay") else "Contrato local"
    return item


def aplicar_metricas_registro(registro, metrics, atualizado_em=None):
    if not metrics:
        return normalizar_registro_cache(registro)

    item = normalizar_registro_cache(registro)
    contrato = metrics.get("contrato") or {}
    calculos = metrics.get("calculos") or {}
    normalizado = metrics.get("normalizado") or {}
    cobrancas = normalizado.get("cobrancas") or []
    atualizado_iso = atualizado_em or datetime.now().isoformat(timespec="seconds")

    try:
        from app.validador_matematico import validar_conciliacao

        status_val, notas, bloqueado = validar_conciliacao(calculos, contrato)
    except Exception:
        status_val, notas, bloqueado = "PENDENTE", [], False

    cliente_real = contrato.get("cliente")
    if cliente_real:
        item["cliente"] = cliente_real
    lote_real = contrato.get("lote")
    if lote_real and item.get("lote") in ("", "-", None):
        item["lote"] = lote_real
    quadra_real = contrato.get("quadra")
    if quadra_real:
        item["quadra"] = quadra_real

    total_parcelas = inteiro(calculos.get("total_parcelas_contrato") or contrato.get("total_parcelas"))
    parcelas_pagas = inteiro(calculos.get("parcelas_pagas_equivalentes"))
    parcelas_restantes = inteiro(calculos.get("parcelas_restantes"))
    valor_total_pago = round(decimal(calculos.get("total_pago_consolidado")), 2)

    item["status"] = status_val
    if total_parcelas > 0 or parcelas_pagas > 0:
        item["parcelas_pagas_identificadas"] = parcelas_pagas
    if total_parcelas > 0:
        item["parcelas_total_contrato"] = total_parcelas
        item["parcelas_restantes"] = parcelas_restantes
    if valor_total_pago > 0:
        item["valor_total_pago"] = valor_total_pago
    item["ultima_atualizacao_widepay"] = atualizado_iso
    item["data_atualizacao"] = atualizado_iso
    aplicar_se_preenchido(item, "valor_base_parcela", contrato.get("valor_parcela"))
    aplicar_se_preenchido(item, "valor_total_contratado", contrato.get("valor_total_contrato"))
    aplicar_se_preenchido(item, "entrada_valor", contrato.get("entrada"))
    item["origem_dados"] = "Contrato local + WidePay"
    item["contrato_modalidade"] = inferir_modalidade_contrato(
        item.get("contrato_arquivo"),
        item.get("pasta_local"),
        contrato,
    )
    item["situacao_final"] = deduzir_situacao_final(item)
    item["contrato_resumo"] = deduzir_resumo_contrato(item)
    item["parcelas_resumo"] = deduzir_resumo_parcelas(item)

    observacoes = []
    if bloqueado:
        observacoes.append("Relatorio bloqueado")
    if notas:
        observacoes.extend(notas[:2])
    if observacoes:
        item["observacoes"] = " | ".join(observacoes)
    elif not item.get("observacoes"):
        item["observacoes"] = "Relatorio gerado"
    if len(notas) > 2:
        item["divergencias"] = " | ".join(notas[2:])

    pagamentos = calculos.get("pagamentos_interpretados") or []
    ultimo_pgto = None
    ultima_data = None
    for pagamento in pagamentos:
        data_ref = pagamento.get("pagamento") or pagamento.get("vencimento")
        if not data_ref:
            continue
        try:
            dt = datetime.strptime(str(data_ref), "%d/%m/%Y")
        except ValueError:
            continue
        if ultima_data is None or dt > ultima_data:
            ultima_data = dt
            ultimo_pgto = pagamento

    if ultimo_pgto:
        item["ultimo_vencimento_pago"] = ultimo_pgto.get("vencimento", "")
        item["ultima_parcela_paga"] = ultimo_pgto.get("descricao", "")
        valor_ultimo = ultimo_pgto.get("valor_recebido")
        if valor_ultimo is not None:
            item["valor_ultimo_pagamento"] = formatar_moeda(valor_ultimo)

    # Contagem de boletos vencidos no WidePay com abatimento/compensação por avulsos de atraso recebidos
    vencidos_list = [cob for cob in cobrancas if str(cob.get("status", "")).lower() == "vencido"]
    
    compensacoes = 0
    for cob in cobrancas:
        status_lower = str(cob.get("status", "")).lower()
        # Se o boleto avulso foi pago (recebido) e indica pagamento de atrasos
        if status_lower == "recebido" and cob.get("classificacao") == "avulso":
            desc_lower = str(cob.get("descricao", "")).lower()
            if any(term in desc_lower for term in ["atraso", "atrazos", "ref", "referente"]):
                compensacoes += 1
                
    boletos_atrasados = max(0, len(vencidos_list) - compensacoes)
    item["boletos_atrasados"] = boletos_atrasados
    atraso_recente = calcular_atraso_recente(metrics)
    item["status_atraso_qtd"] = atraso_recente
    item["status_atraso_cor"] = classificar_status_atraso(atraso_recente)
    item["status_atraso_rotulo"] = rotulo_status_atraso(atraso_recente)
    item["status_atraso_origem"] = "WidePay"
    item["boletos_atrasados"] = atraso_recente
    item["pagamentos_recentes_5m"] = calcular_pagamentos_recentes_5m(cobrancas)

    return item


def mesclar_registro_indexado(
    pasta,
    lote,
    quadra,
    contrato_status,
    contrato_path,
    origem,
    widepay_status,
    agora,
    registro_anterior=None,
    cliente_padrao="",
    dados_contrato=None,
):
    dados_contrato = dados_contrato or {}
    if registro_anterior:
        registro = dict(registro_anterior)
    else:
        registro = {}

    for campo in PRESERVAR_EM_REINDEXACAO:
        if registro_anterior and campo in registro_anterior:
            registro[campo] = registro_anterior.get(campo)

    registro["selecionado"] = bool(registro.get("selecionado", False))
    registro["cliente"] = cliente_padrao or registro.get("cliente") or ""
    lote_contrato = dados_contrato.get("lote")
    quadra_contrato = dados_contrato.get("quadra")
    registro["lote"] = lote_contrato if valor_preenchido(lote_contrato) else lote
    registro["quadra"] = quadra_contrato if valor_preenchido(quadra_contrato) else quadra
    if (
        registro_anterior
        and registro_anterior.get("contrato") == "Encontrado"
        and contrato_status != "Encontrado"
    ):
        contrato_status = "Encontrado"
        contrato_path = registro_anterior.get("contrato_arquivo") or contrato_path
        if not dados_contrato and contrato_path:
            dados_contrato = ler_dados_contrato_arquivo(contrato_path, registro.get("cliente"))
    registro["contrato"] = contrato_status
    registro["contrato_arquivo"] = contrato_path
    registro["origem"] = origem
    registro["pasta_local"] = str(pasta)
    registro["contrato_modalidade"] = inferir_modalidade_contrato(contrato_path, str(pasta), dados_contrato)
    aplicar_se_preenchido(registro, "parcelas_total_contrato", dados_contrato.get("total_parcelas"))
    aplicar_se_preenchido(registro, "valor_base_parcela", dados_contrato.get("valor_parcela"))
    aplicar_se_preenchido(registro, "valor_total_contratado", dados_contrato.get("valor_total_contrato"))
    aplicar_se_preenchido(registro, "entrada_valor", dados_contrato.get("entrada"))
    if dados_contrato:
        registro["origem_contrato"] = "Contrato local"
        registro["origem_dados"] = "Contrato local + cache preservado"

    if not registro.get("status"):
        registro["status"] = "Pendente validacao WidePay" if contrato_status == "Encontrado" else "Sem contrato confirmado"
    elif contrato_status != "Encontrado" and registro.get("status") == "Pendente validacao WidePay":
        registro["status"] = "Sem contrato confirmado"

    if registro.get("observacoes") in (None, ""):
        registro["observacoes"] = widepay_status

    registro["data_atualizacao"] = agora.isoformat(timespec="seconds")

    return normalizar_registro_cache(registro)


def iterar_pastas_cliente():
    base = config.CONTRATOS_DIR
    if not base.exists():
        return []
    candidatos = []
    for path in base.rglob("*"):
        if not path.is_dir():
            continue
        try:
            partes_relativas = path.relative_to(base).parts
        except ValueError:
            partes_relativas = path.parts
        if any(PASTA_IGNORADA_RE.search(normalizar(parte).lower()) for parte in partes_relativas):
            continue
        nome = path.name.strip()
        if not nome or nome.lower().startswith(("_", ".")) or "backup" in nome.lower():
            continue
        lote = extrair_lote(nome)
        contrato_status, contrato_path = detectar_contrato(path)
        if lote != "-" or contrato_status == "Encontrado":
            candidatos.append((path, lote, contrato_status, contrato_path))
    return candidatos


def obter_nome_real_contrato(contrato_path, nome_sugerido):
    if not contrato_path:
        return nome_sugerido
    try:
        p = Path(contrato_path)
        ext = p.suffix.lower()
        texto = ""
        if p.exists():
            if ext == '.docx':
                texto = extrair_texto_docx(p)
            elif ext == '.pdf':
                texto = extrair_texto_pdf(p)
        
        # Fallback: se o arquivo não existe ou é um scan ilegível, busca na pasta 01_DOCUMENTOS_CONVERTIDOS
        if not texto:
            convertidos_dir = config.ROOT_DIR / "01_DOCUMENTOS_CONVERTIDOS"
            if convertidos_dir.exists():
                slug_sug = slug_busca(nome_sugerido)
                for txt_path in convertidos_dir.glob("*.txt"):
                    if slug_sug in slug_busca(txt_path.stem):
                        try:
                            with open(txt_path, "r", encoding="utf-8") as tf:
                                texto = tf.read()
                            if texto:
                                break
                        except Exception:
                            pass
        
        if texto:
            return extrair_nome_cliente(texto, nome_sugerido)
    except Exception:
        pass
    return nome_sugerido


def ler_dados_contrato_arquivo(contrato_path, nome_sugerido):
    dados = {}
    if not contrato_path:
        return dados
    try:
        p = Path(contrato_path)
        texto = ""
        if p.exists():
            ext = p.suffix.lower()
            if ext == ".docx":
                texto = extrair_texto_docx(p)
            elif ext == ".pdf":
                texto = extrair_texto_pdf(p)
            elif ext == ".txt":
                texto = p.read_text(encoding="utf-8", errors="ignore")

        if not texto:
            convertidos_dir = config.ROOT_DIR / "01_DOCUMENTOS_CONVERTIDOS"
            if convertidos_dir.exists():
                slug_sug = slug_busca(nome_sugerido)
                for txt_path in convertidos_dir.glob("*.txt"):
                    if slug_sug in slug_busca(txt_path.stem):
                        texto = txt_path.read_text(encoding="utf-8", errors="ignore")
                        if texto:
                            break

        if texto:
            dados = parsear_dados_contrato(texto, nome_sugerido) or {}
    except Exception:
        return {}
    return dados


def carregar_metricas_historicas_locais(cliente, lote):
    """Busca no diretório de entrega o relatório JSON de métricas mais recente do cliente, usando busca flexível."""
    try:
        import re
        from app.rastreabilidade import re_slug
        cli_slug = re_slug(cliente)
        lot_slug = re_slug(lote)
        
        primeira_palavra = cli_slug.split("_")[0]
        
        # Limpar o lote (ex: "05", "5")
        num_lote = re.sub(r"[^0-9]", "", lot_slug)
        
        pasta_candidata = None
        if config.OUTPUT_DIR.exists():
            for p in config.OUTPUT_DIR.iterdir():
                if not p.is_dir() or not p.name.endswith("_FINAL"):
                    continue
                name_upper = p.name.upper()
                # A pasta deve conter a primeira palavra do cliente
                if primeira_palavra in name_upper:
                    # E o lote deve coincidir em slug completo ou parte numérica
                    if lot_slug in name_upper or (num_lote and num_lote.lstrip("0") in name_upper):
                        # A pasta deve conter de fato arquivos de métricas
                        if list(p.glob("METRICAS_*.json")):
                            pasta_candidata = p
                            break
                        
        if not pasta_candidata:
            pasta_candidata = config.OUTPUT_DIR / f"{cli_slug}_LOTE_{lot_slug}_FINAL"
            
        if not pasta_candidata.exists():
            return None, None
            
        arquivos = list(pasta_candidata.glob("METRICAS_*.json"))
        if not arquivos:
            return None, None
            
        # Pegar o arquivo mais recentemente modificado
        arquivos.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        mais_recente = arquivos[0]
        
        with open(mais_recente, "r", encoding="utf-8") as f:
            metrics_data = json.load(f)
            
        mtime_iso = datetime.fromtimestamp(mais_recente.stat().st_mtime).isoformat(timespec="seconds")
        return metrics_data, mtime_iso
    except Exception:
        return None, None


def chave_registro_cache(registro):
    pasta = str(registro.get("pasta_local") or "").strip().lower()
    lote = chave_lote_canonica(registro) or str(registro.get("lote") or "-").strip().upper()
    cliente = slug_busca(registro.get("cliente") or "")
    if pasta:
        return ("pasta", pasta, lote)
    return ("cliente_lote", cliente, lote)


def pontuar_completude(registro):
    pontos = 0
    if registro.get("contrato") == "Encontrado":
        pontos += 100
    if valor_preenchido(registro.get("cliente")):
        pontos += 15
    if chave_lote_canonica(registro):
        pontos += 15
    if registro.get("contrato_modalidade") not in ("", "nao_confirmado", "NAO CONFIRMADO", None):
        pontos += 10
    for campo in (
        "valor_total_contratado",
        "valor_total_pago",
        "parcelas_total_contrato",
        "parcelas_pagas_identificadas",
        "ultima_atualizacao_widepay",
        "contrato_arquivo",
    ):
        if valor_preenchido(registro.get(campo)):
            pontos += 5
    return pontos


def mesclar_duplicados(principal, candidato):
    if pontuar_completude(candidato) > pontuar_completude(principal):
        principal, candidato = dict(candidato), principal
    else:
        principal = dict(principal)

    for campo, valor in candidato.items():
        if campo in ("observacoes", "divergencias") and valor_preenchido(valor):
            atual = principal.get(campo)
            if valor and valor != atual:
                principal[campo] = f"{atual} | {valor}" if atual else valor
            continue
        if not valor_preenchido(principal.get(campo)) and valor_preenchido(valor):
            principal[campo] = valor

    chave_can = chave_lote_canonica(principal) or chave_lote_canonica(candidato)
    if chave_can:
        principal["chave_lote_canonica"] = chave_can
        principal["lote"] = chave_can
        principal["quadra"] = chave_can[:1]
    return normalizar_registro_cache(principal)


def deduplicar_preservando_ordem(registros):
    unicos = {}
    ordem = []
    for registro in registros:
        chave = chave_registro_cache(registro)
        if chave not in unicos:
            ordem.append(chave)
            unicos[chave] = registro
        else:
            unicos[chave] = mesclar_duplicados(unicos[chave], registro)
    return [unicos[chave] for chave in ordem]


def montar_metricas_resumo_widepay(dados_raw_wp, dados_contrato):
    from app.calculadora_financeira import calcular_resumo
    from app.normalizador_pagamentos import normalizar_e_deduplicar

    valor_base = float(dados_contrato.get("valor_parcela") or 0.0)
    dados_normalizados = normalizar_e_deduplicar(dados_raw_wp, valor_base)
    dados_calculados = calcular_resumo(dados_normalizados, dados_contrato)
    return {
        "contrato": dados_contrato,
        "calculos": dados_calculados,
        "normalizado": dados_normalizados,
    }


def atualizar_resumo_widepay_incremental(registros, log_callback=None, progress_callback=None):
    confirmados = [r for r in registros if r.get("contrato") == "Encontrado"]
    if not confirmados:
        return registros

    def log(msg):
        if log_callback:
            log_callback(msg)

    try:
        import asyncio
        from app.extrator_widepay import extrair_dados_clientes_bloco
        from app.login_navegador import garantir_navegador_conectado

        ws_url = garantir_navegador_conectado()

        async def _coletar():
            atualizados_por_chave = {}
            payload = []
            contrato_por_nome = {}
            for reg in confirmados:
                dados_contrato = {
                    "cliente": reg.get("cliente"),
                    "lote": reg.get("lote"),
                    "quadra": reg.get("quadra"),
                    "valor_parcela": decimal(reg.get("valor_base_parcela")),
                    "total_parcelas": inteiro(reg.get("parcelas_total_contrato")),
                    "valor_total_contrato": decimal(reg.get("valor_total_contratado")),
                    "entrada": decimal(reg.get("entrada_valor")),
                }
                nome = dados_contrato["cliente"] or reg.get("cliente") or ""
                contrato_por_nome[nome] = (reg, dados_contrato)
                contrato_por_nome[normalizar(nome).lower()] = (reg, dados_contrato)
                payload.append({
                    "nome": nome,
                    "lote": dados_contrato.get("lote") or "-",
                    "quadra": dados_contrato.get("quadra") or "-",
                })
            
            inicio_coleta = datetime.now().isoformat(timespec="seconds")
            log(f"WidePay global: coletando boletos/carnes de {len(payload)} cliente(s) em uma varredura paginada.")
            try:
                resultado_bloco = await extrair_dados_clientes_bloco(ws_url, payload, progress_callback=progress_callback)
            except Exception as e:
                log(f"Erro na extracao em lote: {e}")
                resultado_bloco = {}
            boletos_cache = registros_de_resultado_bloco(resultado_bloco)
            metadados_cache = {
                "inicio_coleta": inicio_coleta,
                "fim_coleta": datetime.now().isoformat(timespec="seconds"),
                "total_paginas": "",
                "total_registros_coletados": len(boletos_cache),
                "total_carnes": sum(1 for b in boletos_cache if b.get("fonte") == "carne"),
                "total_cobrancas": sum(1 for b in boletos_cache if b.get("fonte") == "cobranca"),
                "total_clientes_reconhecidos": len({b.get("cliente_normalizado") for b in boletos_cache if b.get("cliente_normalizado")}),
                "origem": "Atualizar clientes / coleta global WidePay",
            }
            if progress_callback:
                progress_callback("Salvando Cache", 70, "Salvando banco interno WidePay...")
            salvar_cache_boletos_widepay(boletos_cache, metadados_cache)
            log(f"Banco interno WidePay salvo: {len(boletos_cache)} boleto(s)/carne(s) em {config.WIDEPAY_BOLETOS_CACHE_JSON}.")
                
            for nome_achado, dados_raw in (resultado_bloco or {}).items():
                reg, dados_contrato = contrato_por_nome.get(nome_achado, (None, None))
                if not reg:
                    reg, dados_contrato = contrato_por_nome.get(normalizar(nome_achado).lower(), (None, None))
                if not reg:
                    continue
                metrics = montar_metricas_resumo_widepay(dados_raw, dados_contrato)
                atualizado = aplicar_metricas_registro(reg, metrics)
                atualizado["origem_dados"] = "Contrato local + WidePay global"
                atualizados_por_chave[chave_registro_cache(reg)] = atualizado
            return atualizados_por_chave

        try:
            atualizados = asyncio.run(_coletar())
        except RuntimeError:
            loop = asyncio.new_event_loop()
            try:
                atualizados = loop.run_until_complete(_coletar())
            finally:
                loop.close()

        return [atualizados.get(chave_registro_cache(r), r) for r in registros]
    except Exception as exc:
        log(f"Atualizacao WidePay falhou: {exc}. Mantendo dados anteriores.")
        mantidos = []
        agora = datetime.now().isoformat(timespec="seconds")
        for reg in registros:
            item = normalizar_registro_cache(reg)
            if item.get("contrato") == "Encontrado":
                obs = item.get("observacoes") or ""
                aviso = "Atualizacao falhou - mantidos dados anteriores"
                item["observacoes"] = aviso if not obs else f"{obs} | {aviso}"
                if not item.get("ultima_atualizacao_widepay"):
                    item["data_atualizacao"] = item.get("data_atualizacao") or agora
            mantidos.append(item)
        return mantidos


def indexar_clientes(validar_widepay=False, log_callback=None, progress_callback=None):
    config.ensure_dirs()
    agora = datetime.now()
    log_path = config.LOG_DIR / f"atualizacao_clientes_{agora.strftime('%Y%m%d_%H%M%S')}.log"

    def log(msg):
        if log_callback:
            log_callback(msg)
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().isoformat(timespec='seconds')}] {msg}\n")

    if progress_callback:
        progress_callback("Preparação", 0, "Preparando atualização...")

    log("Indexacao iniciada")
    widepay_status = "Nao validado nesta atualizacao"
    if progress_callback:
        progress_callback("Validação", 10, "Validando ambiente e conexão com WidePay...")

    if validar_widepay:
        try:
            from app.login_navegador import garantir_navegador_conectado

            garantir_navegador_conectado()
            widepay_status = "CDP WidePay acessivel"
            log("WidePay/CDP validado")
        except Exception as exc:
            widepay_status = f"WidePay indisponivel: {exc}"
            log(widepay_status)

    if progress_callback:
        progress_callback("Leitura Local", 20, "Lendo contratos e base local...")

    registros = []
    vistos = set()
    
    # Carregar cache existente para preservar nomes de clientes já validados pela WidePay ou contratos
    cache_anterior = {}
    try:
        for r in carregar_cache():
            chave_cache = (str(r.get("pasta_local")).lower(), chave_lote_canonica(r) or r.get("lote"))
            cache_anterior[chave_cache] = r
    except Exception:
        pass

    for pasta, lote, contrato_status, contrato_path in iterar_pastas_cliente():
        cliente_sugerido = limpar_nome_cliente(pasta.name)
        cliente = cliente_sugerido
        
        # O nome exibido deve priorizar o contrato local confirmado. Os dados financeiros
        # do cache anterior seguem preservados por mesclar_registro_indexado().
        lote_canonico = normalizar_lote_quadra(lote, "", pasta.name, str(pasta)) or lote
        chave_res = (str(pasta).lower(), lote_canonico)
        r_antigo = cache_anterior.get(chave_res)
        dados_contrato = {}
        if contrato_status == "Encontrado" and contrato_path:
            dados_contrato = ler_dados_contrato_arquivo(contrato_path, cliente_sugerido)
            cliente = dados_contrato.get("cliente") or obter_nome_real_contrato(contrato_path, cliente_sugerido)
        elif r_antigo and r_antigo.get("cliente"):
            cliente = r_antigo["cliente"]
        quadra = extrair_quadra(pasta.name, lote)
        chave = (slug_busca(cliente), lote_canonico if lote_canonico != "-" else str(pasta).lower())
        if chave in vistos:
            continue
        vistos.add(chave)
        origem = "Contrato local"
        registro = mesclar_registro_indexado(
            pasta=pasta,
            lote=lote,
            quadra=quadra,
            contrato_status=contrato_status,
            contrato_path=contrato_path,
            origem=origem,
            widepay_status=widepay_status,
            agora=agora,
            registro_anterior=r_antigo,
            cliente_padrao=cliente,
            dados_contrato=dados_contrato,
        )
        
        # Restauração automática de histórico físico
        if registro.get("contrato") == "Encontrado" and (registro.get("boletos_atrasados") == "" or registro.get("status") == "Pendente validacao WidePay"):
            metrics_data, mtime_iso = carregar_metricas_historicas_locais(cliente, lote)
            if metrics_data:
                registro = aplicar_metricas_registro(registro, metrics_data, atualizado_em=mtime_iso)

        registros.append(registro)

    chaves_atuais = {chave_registro_cache(r) for r in registros}
    for antigo in carregar_cache():
        chave_antiga = chave_registro_cache(antigo)
        if chave_antiga in chaves_atuais:
            continue
        preservado = normalizar_registro_cache(antigo)
        obs = preservado.get("observacoes") or ""
        aviso = "Nao localizado nesta varredura - mantido do cache anterior"
        preservado["observacoes"] = aviso if not obs else f"{obs} | {aviso}"
        registros.append(preservado)
        chaves_atuais.add(chave_antiga)

    registros = deduplicar_preservando_ordem(registros)
    if validar_widepay:
        if progress_callback:
            progress_callback("Conexão WidePay", 30, "Conectando ao WidePay...")
        registros = atualizar_resumo_widepay_incremental(registros, log_callback=log, progress_callback=progress_callback)

    if progress_callback:
        progress_callback("Atualização de Lista", 80, "Atualizando lista de clientes...")
    registros, resumo_saneamento = saneamento_clientes.aplicar_saneamento(registros)
    registros.sort(key=lambda r: (r.get("saneamento_categoria", "ativos"), slug_busca(r["cliente"]), r["lote"]))
    
    if progress_callback:
        progress_callback("Geração XLSX", 90, "Gerando banco de dados XLSX...")
    salvar_cache(registros)
    log(f"Indexacao concluida: {len(registros)} cliente/lote")
    log(
        "Saneamento aplicado: "
        f"{resumo_saneamento.get('ativos', 0)} ativos, "
        f"{resumo_saneamento.get('quitados', 0)} quitados, "
        f"{resumo_saneamento.get('bloqueados_removidos', 0)} bloqueados/removidos, "
        f"{resumo_saneamento.get('atencao', 0)} atencao"
    )
    if progress_callback:
        progress_callback("Concluído", 100, f"Atualização concluída com sucesso. {len(registros)} registros indexados.")
    return {"registros": registros, "log": str(log_path), "widepay_status": widepay_status}


def salvar_cache(registros):
    config.ensure_dirs()
    registros_normalizados = [normalizar_registro_cache(item) for item in registros]
    registros_normalizados, resumo_saneamento = saneamento_clientes.aplicar_saneamento(registros_normalizados)
    payload = {
        "gerado_em": datetime.now().isoformat(timespec="seconds"),
        "quantidade": len(registros_normalizados),
        "resumo_saneamento": resumo_saneamento,
        "registros": registros_normalizados,
    }
    with open(config.CLIENTES_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    salvar_xlsx(registros_normalizados)


def salvar_xlsx(registros):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ativos"
    headers = [
        "cliente",
        "lote",
        "quadra",
        "chave_lote_canonica",
        "status",
        "contrato",
        "contrato_resumo",
        "contrato_modalidade",
        "valor_base_parcela",
        "valor_total_contratado",
        "entrada_valor",
        "origem_contrato",
        "origem",
        "parcelas_pagas_identificadas",
        "parcelas_total_contrato",
        "parcelas_restantes",
        "parcelas_resumo",
        "valor_total_pago",
        "ultima_parcela_paga",
        "ultimo_vencimento_pago",
        "valor_ultimo_pagamento",
        "data_atualizacao",
        "ultima_atualizacao_widepay",
        "situacao_final",
        "status_atraso_qtd",
        "status_atraso_cor",
        "status_atraso_rotulo",
        "status_atraso_origem",
        "observacoes",
        "divergencias",
        "origem_dados",
        "saneamento_acao",
        "saneamento_categoria",
        "saneamento_origem",
        "saneamento_observacao",
        "cliente_canonico",
        "chave_saneamento",
        "quitado_manual",
        "bloqueado_removido_manual",
        "atencao_manual",
        "ignorado_manual",
        "ativo_na_lista_principal",
        "status_operacional",
        "origem_classificacao",
        "bloqueado_reaparecer",
        "motivo_remocao_lista",
        "pasta_local",
    ]

    grupos = saneamento_clientes.separar_por_categoria(registros)

    def preparar_planilha(sheet, titulo, itens):
        sheet.title = titulo
        sheet.append(headers)
        header_fill = PatternFill("solid", fgColor="1C2B33")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for item in itens:
            sheet.append([item.get(h, "") for h in headers])
            if item.get("saneamento_acao") == "atencao":
                for cell in sheet[sheet.max_row]:
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
        for col in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in col) + 2, 60)
            sheet.column_dimensions[col[0].column_letter].width = width

    preparar_planilha(ws, "Ativos", grupos["ativos"])
    pagamentos_recentes = wb.create_sheet("Pagamentos_Recentes")
    recentes_headers = [label for _key, label in COLUNAS_XLSX_PAGAMENTOS_RECENTES]
    meses_recentes = janela_pagamentos_recentes()
    pagamentos_recentes.append(recentes_headers + [mes["rotulo"] for mes in meses_recentes])
    header_fill = PatternFill("solid", fgColor="1C2B33")
    for cell in pagamentos_recentes[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for item in grupos["ativos"]:
        linha_base = [
            limpar_nome_cliente(str(item.get("cliente") or "")),
            f"{(item.get('chave_lote_canonica') or item.get('lote') or '-')} / {((item.get('chave_lote_canonica') or '')[:1] or item.get('quadra') or '-')}",
            deduzir_resumo_contrato(item),
            deduzir_resumo_parcelas(item),
            item.get("situacao_final") or deduzir_situacao_final(item),
            formatar_data_hora(item.get("ultima_atualizacao_widepay") or item.get("data_atualizacao")),
            formatar_moeda(item.get("valor_total_contratado")),
            formatar_moeda(item.get("valor_total_pago")),
        ]
        mapa_recentes = normalizar_pagamentos_recentes_5m(item.get("pagamentos_recentes_5m"))
        linha_meses = []
        for mes in meses_recentes:
            val_mes = mapa_recentes.get(mes["chave"], "-")
            if isinstance(val_mes, dict):
                val_mes = val_mes.get("status") or "-"
            linha_meses.append(val_mes)
        pagamentos_recentes.append(linha_base + linha_meses)
    for col in pagamentos_recentes.columns:
        width = min(max(len(str(cell.value or "")) for cell in col) + 2, 24)
        pagamentos_recentes.column_dimensions[col[0].column_letter].width = width
    preparar_planilha(wb.create_sheet("Quitados"), "Quitados", grupos["quitados"])
    preparar_planilha(wb.create_sheet("Bloqueados_Removidos"), "Bloqueados_Removidos", grupos["bloqueados_removidos"])

    auditoria = wb.create_sheet("Auditoria")
    auditoria.append(["item", "valor"])
    for cell in auditoria[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1C2B33")
    auditoria_linhas = [
        ("gerado_em", datetime.now().isoformat(timespec="seconds")),
        ("total_registros_preservados", len(registros)),
        ("ativos", len(grupos["ativos"])),
        ("quitados", len(grupos["quitados"])),
        ("bloqueados_removidos", len(grupos["bloqueados_removidos"])),
        ("ignorados", len(grupos["ignorados"])),
        ("fora_roster_planilha", len(grupos["fora_roster"])),
        ("amarelos_atencao", len(grupos["atencao"])),
        ("movimentos_manuais", len(saneamento_clientes.carregar_saneamento().get("auditoria_manual", []))),
        ("arquivo_saneamento", str(saneamento_clientes.ARQUIVO_SANEAMENTO)),
        ("planilha_marcada", str(saneamento_clientes.encontrar_planilha_marcada() or "")),
    ]
    for linha in auditoria_linhas:
        auditoria.append(list(linha))
    if grupos["ignorados"]:
        auditoria.append([])
        auditoria.append(["ignorados_cliente", "ignorados_lote"])
        for item in grupos["ignorados"]:
            auditoria.append(
                [
                    item.get("cliente") or "",
                    item.get("chave_lote_canonica") or item.get("lote") or "",
                ]
            )
    if grupos["fora_roster"]:
        auditoria.append([])
        auditoria.append(["fora_roster_cliente", "fora_roster_lote"])
        for item in grupos["fora_roster"]:
            auditoria.append(
                [
                    item.get("cliente") or "",
                    item.get("chave_lote_canonica") or item.get("lote") or "",
                ]
            )
    movimentos = saneamento_clientes.carregar_saneamento().get("auditoria_manual", [])
    if movimentos:
        auditoria.append([])
        auditoria.append(["data_hora", "acao_executada"])
        for mov in movimentos[-50:]:
            auditoria.append(
                [
                    f"{mov.get('data_hora')} | {mov.get('cliente')} | {mov.get('chave_lote_canonica') or mov.get('lote')}",
                    f"{mov.get('acao_executada')} | {mov.get('aba_origem')} -> {mov.get('aba_destino')}",
                ]
            )
    auditoria.column_dimensions["A"].width = 32
    auditoria.column_dimensions["B"].width = 90
        
    # Salvar backup se o arquivo já existir
    if config.VISUAL_XLSX.exists():
        import shutil
        from app import paths
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"BANCO_DADOS_WIDEAPP_EXTRA_ANTES_{timestamp}.xlsx"
        backup_path = paths.get_backups_dir() / backup_name
        try:
            shutil.copy2(config.VISUAL_XLSX, backup_path)
            print(f"Backup do banco visual criado em: {backup_path}")
        except Exception as exc:
            print(f"Erro ao criar backup do banco visual: {exc}")
            
    # Salvar banco visual (raiz do app)
    wb.save(config.VISUAL_XLSX)
    print(f"Banco visual XLSX salvo em: {config.VISUAL_XLSX}")
    
    # Manter cópia na pasta data interna para compatibilidade
    wb.save(config.CLIENTES_XLSX)


def carregar_cache():
    if not config.CLIENTES_JSON.exists():
        return []
    with open(config.CLIENTES_JSON, "r", encoding="utf-8") as f:
        payload = json.load(f)
    registros = []
    modificado = False
    for item in payload.get("registros", []):
        registro = normalizar_registro_cache(item)
        mapa = registro.get("pagamentos_recentes_5m", {})
        is_rico = isinstance(mapa, dict) and mapa.get("_enriquecido") is True and len(mapa) >= 10
        if is_rico:
            # Se for rico mas todos os meses forem "Sem boleto" mesmo o contrato sendo "Encontrado",
            # significa que o cache foi envenenado na geracao antiga. Forcamos re-enriquecimento.
            valores_meses = []
            for k, v in mapa.items():
                if k == "_enriquecido":
                    continue
                if isinstance(v, dict):
                    valores_meses.append(v.get("status"))
                else:
                    valores_meses.append(str(v).strip())
            if len(valores_meses) > 0 and all(v == "Sem boleto" for v in valores_meses):
                if registro.get("contrato") == "Encontrado":
                    is_rico = False
                    
        if not is_rico:
            registro = enriquecer_pagamentos_recentes_cache(registro)
            modificado = True
        registros.append(registro)
        
    registros, resumo_saneamento = saneamento_clientes.aplicar_saneamento(registros)
    
    if modificado:
        payload["registros"] = registros
        payload["resumo_saneamento"] = resumo_saneamento
        payload["gerado_em"] = datetime.now().isoformat(timespec="seconds")
        try:
            with open(config.CLIENTES_JSON, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
            print("Cache de clientes enriquecido para 10 meses e salvo com sucesso.")
        except Exception as e:
            print(f"Erro ao salvar cache enriquecido: {e}")
            
    return registros
