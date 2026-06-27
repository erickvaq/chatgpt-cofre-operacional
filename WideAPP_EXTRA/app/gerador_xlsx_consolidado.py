# -*- coding: utf-8 -*-
"""Gerador de XLSX consolidado para selecoes da interface."""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app import config


ABAS = [
    "Clientes processados",
    "Resumo financeiro",
    "Pagamentos confirmados",
    "Parcelas pagas",
    "Parcelas restantes",
    "Atrasos",
    "Boletos avulsos",
    "Entradas",
    "Divergencias",
    "Status final",
]


def gerar(registros, grupo="SELECIONADOS"):
    config.ensure_dirs()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho = config.OUTPUT_DIR / f"CONSOLIDADO_WIDEAPP_EXTRA_{grupo}_{stamp}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = ABAS[0]
    for aba in ABAS[1:]:
        wb.create_sheet(aba)

    headers = [
        "cliente",
        "lote",
        "quadra",
        "status",
        "contrato",
        "contrato_resumo",
        "parcelas_resumo",
        "situacao_final",
        "ultima_atualizacao_widepay",
        "valor_total_pago",
        "parcelas_pagas_identificadas",
        "parcelas_total_contrato",
        "parcelas_restantes",
        "ultima_parcela_paga",
        "ultimo_vencimento_pago",
        "valor_ultimo_pagamento",
        "observacoes",
        "divergencias",
    ]
    for sheet in wb.worksheets:
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        if sheet.title == "Clientes processados":
            for item in registros:
                sheet.append([item.get(h, "") for h in headers])
        else:
            sheet.append(["Pendente de extracao WidePay individual pelo pipeline financeiro"] + [""] * (len(headers) - 1))
        for col in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in col) + 2, 60)
            sheet.column_dimensions[col[0].column_letter].width = width

    wb.save(caminho)
    return caminho
